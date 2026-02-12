"""
Statistics API Routes - V1

This module contains all statistics-related routes:
- GET /api/stats/top - Download statistics with rankings
- GET /api/stats/export - Export statistics as Excel file
- GET /stats - Statistics page (public)
- GET /api/admin/stats - Admin statistics (admin only)
"""

import io
import logging
from datetime import date
from typing import Dict, List, Optional, Any

from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl import Workbook
from starlette.requests import Request as StarletteRequest

# Templates (will be injected from main app)
templates = None

# Logger
logger = logging.getLogger("skillhub")

# Router
router = APIRouter(prefix="/api/v1", tags=["stats"])

# Dependencies from database module
from database import (
    get_download_stats,
    get_stats_with_author,
    get_total_users_count,
    get_skills_count_by_status,
    get_today_downloads_count,
    get_top_skills_by_downloads,
    get_top_users_by_downloads,
)

# Import helper functions from core/dependencies
from core.dependencies import get_current_user, require_admin

# Import scan_plugins from skills module
from api.v1.skills import scan_plugins


# ============================================================================
# API Routes (Public Statistics)
# ============================================================================

@router.get("/stats/top")
async def api_stats_top(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)")
):
    """Get download statistics with rankings."""
    try:
        # Parse dates
        start = date.fromisoformat(start_date) if start_date else None
        end = date.fromisoformat(end_date) if end_date else None

        # Get plugins for author mapping
        plugins = scan_plugins()

        # Get stats with author info
        stats = get_stats_with_author(plugins, start, end)

        return {
            "period": {
                "start_date": start_date or "all-time",
                "end_date": end_date or "all-time"
            },
            "total_downloads": stats["total_downloads"],
            "rankings": stats["rankings"]
        }

    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid date format: {e}"
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get stats: {e}"
        )


@router.get("/stats/export")
async def api_stats_export(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)")
):
    """Export download statistics as Excel file."""
    try:
        # Parse dates
        start = date.fromisoformat(start_date) if start_date else None
        end = date.fromisoformat(end_date) if end_date else None

        # Get plugins for author mapping
        plugins = scan_plugins()

        # Get stats with author info
        stats = get_stats_with_author(plugins, start, end)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Download Statistics"

        # Header row
        headers = ["排名", "Skill 名称", "作者", "下载次数"]
        ws.append(headers)

        # Style header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows
        for idx, ranking in enumerate(stats["rankings"], 1):
            ws.append([
                idx,
                ranking["skill_name"],
                ranking["author"],
                ranking["downloads"]
            ])

        # Style data rows
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in range(2, len(stats["rankings"]) + 2):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if col in [1, 4] else "left", vertical="center")

        # Adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename
        period_str = f"{start_date or 'all'}_to_{end_date or 'all'}"
        filename = f"download_stats_{period_str}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel export requires openpyxl: pip install openpyxl"
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid date format: {e}"
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export stats: {e}"
        )


# ============================================================================
# API Routes (Admin Statistics)
# ============================================================================

@router.get("/admin/stats")
async def api_admin_stats(
    _: bool = Depends(require_admin)
):
    """Get admin statistics (admin only).

    Returns comprehensive statistics about the registry including:
    - Total users count
    - Pending skills count
    - Approved skills count
    - Today's downloads count
    - Top 10 skills by downloads
    - Top 10 users by downloads
    """
    try:
        # Get counts
        total_users = get_total_users_count()
        pending_skills = get_skills_count_by_status("pending")
        approved_skills = get_skills_count_by_status("approved")
        today_downloads = get_today_downloads_count()

        # Get top rankings
        top_skills = get_top_skills_by_downloads(10)
        top_users = get_top_users_by_downloads(10)

        return {
            "success": True,
            "data": {
                "counts": {
                    "total_users": total_users,
                    "pending_skills": pending_skills,
                    "approved_skills": approved_skills,
                    "today_downloads": today_downloads
                },
                "top_skills": top_skills,
                "top_users": top_users
            }
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get admin stats: {str(e)}"
        )


# ============================================================================
# UI Routes (Statistics Pages)
# ============================================================================

@router.get("/stats", response_class=HTMLResponse)
async def stats_page(request: StarletteRequest):
    """Display download statistics page (public access)."""
    return templates.TemplateResponse("stats.html", {
        "request": request
    })


# ============================================================================
# Initialization
# ============================================================================

def init_stats_router(templates_instance: Jinja2Templates) -> APIRouter:
    """Initialize and configure the stats router.

    Args:
        templates_instance: Jinja2Templates instance for rendering templates

    Returns:
        Configured APIRouter instance
    """
    global templates
    templates = templates_instance
    return router
