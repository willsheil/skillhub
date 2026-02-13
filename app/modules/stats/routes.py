"""
Stats module API routes.

Provides endpoints for download statistics and rankings.
"""

import io
import logging
from datetime import date
from typing import Optional

from fastapi import APIRouter, HTTPException, Query, Response, Depends
from fastapi.responses import StreamingResponse
from starlette.requests import Request

from app.modules.stats.services import (
    get_stats_with_author,
    get_download_stats,
    get_today_downloads_count,
    get_top_skills_by_downloads,
    get_top_users_by_downloads,
    record_download
)
from app.modules.stats.dependencies import parse_date_range, require_admin

# Import scan_plugins for author mapping
# TODO: This should be moved to a shared service in the future
from main import scan_plugins

logger = logging.getLogger("skillhub.stats.routes")

router = APIRouter(prefix="/api/stats", tags=["stats"])


@router.get("/top")
async def get_stats_top(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)")
):
    """Get download statistics with rankings."""
    try:
        # Parse dates
        start = date.fromisoformat(start_date) if start_date else None
        end = date.fromisoformat(end_date) if end_date else None

        # Get plugins for author mapping
        plugins = scan_plugins()

        # Get stats with author info
        stats = get_stats_with_author(plugins, start, end)

        return {
            "period": {
                "start_date": start_date or "all-time",
                "end_date": end_date or "all-time"
            },
            "total_downloads": stats["total_downloads"],
            "rankings": stats["rankings"]
        }

    except ValueError as e:
        raise HTTPException(400, f"Invalid date format: {e}")
    except Exception as e:
        raise HTTPException(500, f"Failed to get stats: {e}")


@router.get("/export")
async def export_stats(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)")
):
    """Export download statistics as Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # Parse dates
        start = date.fromisoformat(start_date) if start_date else None
        end = date.fromisoformat(end_date) if end_date else None

        # Get plugins for author mapping
        plugins = scan_plugins()

        # Get stats with author info
        stats = get_stats_with_author(plugins, start, end)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Download Statistics"

        # Header row
        headers = ["排名", "Skill 名称", "作者", "下载次数"]
        ws.append(headers)

        # Style header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows
        for idx, ranking in enumerate(stats["rankings"], 1):
            ws.append([
                idx,
                ranking["skill_name"],
                ranking["author"],
                ranking["downloads"]
            ])

        # Style data rows
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in range(2, len(stats["rankings"]) + 2):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if col in [1, 4] else "left", vertical="center")

        # Adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename
        period_str = f"{start_date or 'all'}_to_{end_date or 'all'}"
        filename = f"download_stats_{period_str}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except ImportError:
        raise HTTPException(500, "Excel export requires openpyxl: pip install openpyxl")
    except ValueError as e:
        raise HTTPException(400, f"Invalid date format: {e}")
    except Exception as e:
        raise HTTPException(500, f"Failed to export stats: {e}")


@router.get("/admin")
async def get_admin_stats(
    _: bool = Depends(require_admin)
):
    """Get admin statistics (admin only).

    Returns comprehensive statistics about the registry including:
    - Total users count
    - Pending skills count
    - Approved skills count
    - Today's downloads count
    - Top 10 skills by downloads
    - Top 10 users by downloads
    """
    try:
        # Import these functions from database module
        # TODO: These should be refactored into appropriate services
        from database import (
            get_total_users_count,
            get_skills_count_by_status
        )

        # Get counts
        total_users = get_total_users_count()
        pending_skills = get_skills_count_by_status("pending")
        approved_skills = get_skills_count_by_status("approved")
        today_downloads = get_today_downloads_count()

        # Get top rankings
        top_skills = get_top_skills_by_downloads(10)
        top_users = get_top_users_by_downloads(10)

        return {
            "success": True,
            "data": {
                "total_users": total_users,
                "pending_skills": pending_skills,
                "approved_skills": approved_skills,
                "today_downloads": today_downloads,
                "top_skills": top_skills,
                "top_users": top_users
            }
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to fetch admin statistics: {str(e)}"
        )
