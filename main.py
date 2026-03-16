#!/usr/bin/env python3
"""
Claude Code Skill Registry - Private Marketplace Server
"""

# Load environment variables from .env file
from dotenv import load_dotenv
load_dotenv()

import asyncio
import json
import logging
import os
import re
import secrets
import shutil
from contextlib import asynccontextmanager
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime, date
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.interval import IntervalTrigger

import yaml
import markdown
from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Depends, status, Query
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, Response, PlainTextResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.requests import Request
from starlette.middleware.sessions import SessionMiddleware
from pydantic import BaseModel

# Import and setup logging configuration
from utils.logging_config import setup_logging, audit_log, PerformanceTracker

# Initialize logging system
setup_logging(
    level=os.getenv("LOG_LEVEL", "INFO"),
    log_dir="./logs",
    enable_json=True,
    enable_console=True
)

# Configuration
PLUGINS_DIR = Path(os.getenv("PLUGINS_DIR", "./plugins"))
PLUGINS_DIR.mkdir(exist_ok=True)

# Pending uploads directory
PENDING_DIR = Path("./data/pending")
PENDING_DIR.mkdir(parents=True, exist_ok=True)

# Import database module
from database import (
    init_db, record_download, get_download_stats, get_stats_with_author,
    get_user_by_credentials, get_user_by_id, update_last_login,
    create_skill_record, get_pending_skills, get_skill_by_id,
    update_skill_status, get_user_uploads, get_total_users_count,
    get_skills_count_by_status, get_today_downloads_count,
    get_top_skills_by_downloads, get_top_users_by_downloads, get_upload_stats,
    get_skill_source_type, create_notification, update_skill_active_status,
    get_skill_active_status, get_my_skills,
    get_user_notifications, get_unread_notifications_count,
    mark_notification_read, mark_all_notifications_read, cleanup_old_notifications,
    get_users_list, create_user, update_user_role, disable_user,
    enable_user, delete_user, reset_user_api_key, get_user_skills_count,
    get_skill_approval_status, delete_skill_version, batch_unlist_skills,
    batch_delete_skills,
    get_api_keys_list, create_api_key, delete_api_key, toggle_api_key_status,
    get_api_key_stats
)

# Configure logging - get application logger with proper configuration
logger = logging.getLogger("skillhub")

# Admin credentials (can be overridden via environment variables)
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin123")  # 默认密码，生产环境应修改

# Session secret key (should be changed in production)
SECRET_KEY = os.getenv("SECRET_KEY", "your-secret-key-change-this")

# API key reset rate limiting (user_id -> last_reset_time)
_api_key_reset_times = {}


# Global scheduler instance
scheduler = AsyncIOScheduler()

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Lifespan context manager for startup and shutdown events."""
    # Startup
    init_db()

    # Start Gitea push service if configured
    gitea_url = os.getenv("GITEA_REPO_URL")
    logger.info(f"GITEA_REPO_URL = {gitea_url}")
    if gitea_url:
        try:
            logger.info("Starting Gitea push service with APScheduler...")
            from services.gitea import GiteaPushService

            push_service = GiteaPushService(
                interval=int(os.getenv("GITEA_PUSH_INTERVAL", "30"))
            )

            # Import the sync wrapper
            from services.gitea.gitea_push_service import run_push_task

            # Start APScheduler with interval trigger
            interval_seconds = int(os.getenv("GITEA_PUSH_INTERVAL", "30"))
            scheduler.add_job(
                run_push_task,
                trigger=IntervalTrigger(seconds=interval_seconds),
                id="gitea_push_task",
                name="Gitea Push Service",
                replace_existing=True
            )
            scheduler.start()
            logger.info(f"Gitea push service scheduled every {interval_seconds} seconds")
        except Exception as e:
            import traceback
            logger.error(f"Failed to start Gitea push service: {e}")
            logger.error(traceback.format_exc())
    else:
        logger.info("Gitea integration disabled (GITEA_REPO_URL not set)")

    yield

    # Shutdown
    logger.info("Shutting down...")
    scheduler.shutdown()


app = FastAPI(
    title="SkillHub API",
    description="Claude Code 技能插件管理系统 API",
    version="1.0.0",
    docs_url="/api/v1/docs",
    redoc_url="/api/v1/redoc",
    lifespan=lifespan
)

# Register external API v1 router
from apps import router as api_v1_router
app.include_router(api_v1_router)

# 配置 API Key 安全方案
from fastapi.openapi.utils import get_openapi

def custom_openapi():
    if app.openapi_schema:
        return app.openapi_schema

    openapi_schema = get_openapi(
        title="SkillHub External API",
        version="1.0.0",
        routes=app.routes,
    )

    # 添加安全方案
    openapi_schema["components"]["securitySchemes"] = {
        "ApiKeyAuth": {
            "type": "apiKey",
            "in": "header",
            "name": "X-API-Key",
            "description": "请输入您的 API Key"
        }
    }

    # 全局应用安全方案
    for path in openapi_schema["paths"].values():
        for operation in path.values():
            operation["security"] = [{"ApiKeyAuth": []}]

    app.openapi_schema = openapi_schema
    return app.openapi_schema

app.openapi = custom_openapi

# Add session middleware
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY)

# Static files and templates
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# Set templates for pages router
from apps.pages import set_templates
set_templates(templates)


def require_auth(request: Request):
    """Check if user is authenticated.

    Raises HTTP 401 if user is not logged in.
    """
    user_id = request.session.get("user_id")
    if not user_id:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Not authenticated"
        )
    return True


def require_admin(request: Request):
    """Check if user is authenticated and has admin role.

    Raises HTTP 401 if user is not logged in.
    Raises HTTP 403 if user is not an admin.
    """
    user_id = request.session.get("user_id")
    if not user_id:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Not authenticated"
        )

    role = request.session.get("role")
    if role != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )
    return True


def get_current_user(request: Request) -> Optional[dict]:
    """Get the current authenticated user from session.

    Returns:
        User dictionary if authenticated, None otherwise
    """
    user_id = request.session.get("user_id")
    if not user_id:
        return None

    return get_user_by_id(user_id)


def verify_credentials(username: str, password: str) -> bool:
    """Verify admin credentials."""
    return username == ADMIN_USERNAME and password == ADMIN_PASSWORD


class PluginMetadata(BaseModel):
    name: str
    version: str
    description: str
    author: dict
    updated_at: Optional[str] = None


def get_skill_dir_name(filename: str) -> str:
    """Get the skill directory name from filename.

    Removes .zip extension to get the directory name.
    Example: skill-name-1.0.0.zip -> skill-name-1.0.0
    """
    return filename[:-4] if filename.endswith('.zip') else filename


def parse_skill_md(content: str) -> Tuple[Optional[dict], str]:
    """Parse SKILL.md content to extract YAML frontmatter and markdown body.

    Args:
        content: Raw SKILL.md content

    Returns:
        Tuple of (yaml_metadata_dict, markdown_body)
        If no YAML frontmatter found, returns (None, content)
    """
    # Pattern to match YAML frontmatter between --- markers
    # Use .+? instead of .*? to ensure we match at least one character
    # Remove ^ anchor since content might have leading whitespace
    pattern = r'---\s*\n(.+?)\n---\s*\n(.*)'
    match = re.match(pattern, content, re.DOTALL)

    if match:
        yaml_content = match.group(1)
        markdown_body = match.group(2).strip()
        try:
            metadata = yaml.safe_load(yaml_content)
            if not isinstance(metadata, dict):
                metadata = {}
            return metadata, markdown_body
        except yaml.YAMLError:
            return None, content

    return None, content


def extract_metadata_from_skill_md(zip_path: Path) -> Optional[dict]:
    """Extract metadata from SKILL.md inside zip.

    The ZIP should have structure:
        skill-name/
        └── SKILL.md

    Args:
        zip_path: Path to the skill ZIP file

    Returns:
        Metadata dict or None if parsing fails
    """
    import zipfile

    # 快速检查文件是否存在，避免不必要的异常处理和日志
    if not zip_path.exists():
        return None

    try:
        with zipfile.ZipFile(zip_path, 'r') as zf:
            namelist = zf.namelist()

            # Find SKILL.md (may be in root or subdirectory)
            skill_md_paths = [name for name in namelist
                             if name.endswith('SKILL.md') or name == 'SKILL.md']

            if not skill_md_paths:
                return None

            # Use the first SKILL.md found
            skill_md_path = skill_md_paths[0]
            content = zf.read(skill_md_path).decode('utf-8')

            metadata, _ = parse_skill_md(content)
            return metadata

    except Exception as e:
        logger.debug(f"Failed to extract metadata from {zip_path}: {e}")
        return None


def validate_skill_name(name: str) -> tuple[bool, str]:
    """Validate skill name according to specification.

    Requirements:
    - Must be 1-64 characters
    - May only contain lowercase letters, numbers, and hyphens
    - Must not start or end with '-'
    - Must not contain consecutive hyphens ('--')

    Returns:
        (is_valid, error_message)
    """
    if not name or not isinstance(name, str):
        return False, "Name is required"

    if len(name) < 1 or len(name) > 64:
        return False, "Name must be 1-64 characters"

    if not re.match(r'^[a-z0-9-]+$', name):
        return False, "Name may only contain lowercase letters, numbers, and hyphens"

    if name.startswith('-') or name.endswith('-'):
        return False, "Name must not start or end with hyphen"

    if '--' in name:
        return False, "Name must not contain consecutive hyphens"

    return True, ""


def package_skill_with_installer(
    original_zip_path: Path,
    skill_name: str,
    version: str
) -> bytes:
    """Package skill with installer scripts.

    Reads the original skill ZIP, adds install.bat and install.sh scripts,
    and returns the new ZIP as bytes.

    Args:
        original_zip_path: Path to the original skill ZIP file
        skill_name: Name of the skill
        version: Version of the skill

    Returns:
        ZIP file content as bytes with installer scripts included
    """
    import zipfile
    import io

    # Read script templates
    templates_dir = Path(__file__).parent / "templates" / "install_scripts"

    # Render script templates
    skill_dir_name = get_skill_dir_name(original_zip_path.name)

    script_vars = {
        "skill_name": skill_name,
        "version": version,
        "skill_dir_name": skill_dir_name
    }

    # Read and render install.bat
    bat_template = (templates_dir / "install.bat").read_text(encoding='utf-8')
    bat_content = bat_template
    for key, value in script_vars.items():
        bat_content = bat_content.replace(f"{{{{{key}}}}}", value)

    # Read and render install.sh
    sh_template = (templates_dir / "install.sh").read_text(encoding='utf-8')
    sh_content = sh_template
    for key, value in script_vars.items():
        sh_content = sh_content.replace(f"{{{{{key}}}}}", value)

    # Create new ZIP with installer scripts
    output = io.BytesIO()

    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zf_out:
        # Copy all content from original ZIP
        with zipfile.ZipFile(original_zip_path, 'r') as zf_in:
            for item in zf_in.namelist():
                zf_out.writestr(item, zf_in.read(item))

        # Add installer scripts at root level (BAT with UTF-8 BOM for Windows)
        zf_out.writestr('install.bat', bat_content.encode('utf-8-sig'))
        zf_out.writestr('install.sh', sh_content.encode('utf-8'))

        # Add README
        readme_content = f"""Claude Code Skill: {skill_name} v{version}
{'=' * 50}

一键安装 (推荐):
----------------
Windows:
  1. 解压此 ZIP 文件
  2. 双击运行 install.bat
  3. 按提示完成安装

Linux / macOS:
  1. 解压此 ZIP 文件
  2. 打开终端，cd 到解压目录
  3. 运行: chmod +x install.sh && ./install.sh
  4. 按提示完成安装

手动安装:
---------
1. 将此 ZIP 解压到 Claude Code Skills 目录:
   - Windows: %USERPROFILE%\\.claude\\skills\
   - Linux/Mac: ~/.claude/skills/

2. 重启 Claude Code 以加载新技能

环境变量设置 (可选):
--------------------
设置 CLAUDE_SKILLS_PATH 环境变量可让安装脚本自动识别路径:
  - Windows: setx CLAUDE_SKILLS_PATH "C:\\path\\to\\skills"
  - Linux/Mac: export CLAUDE_SKILLS_PATH="/path/to/skills"

"""
        zf_out.writestr('README.txt', readme_content)

    return output.getvalue()


def parse_plugin_filename(filename: str) -> tuple[str, str]:
    """Parse plugin filename to extract skill name.

    Format: {skill-name}.zip (version is no longer extracted from filename)
    Example: ask-questions-if-underspecified.zip
             semgrep-rule-creator.zip

    Note: Version is now specified in SKILL.md metadata instead of filename.

    Returns: (skill_name, "unknown") - version always returns "unknown" to indicate
             it should be read from SKILL.md metadata
    """
    # Remove .zip extension
    skill_name = filename[:-4] if filename.endswith('.zip') else filename

    # Version is no longer parsed from filename
    # It should be read from SKILL.md metadata
    return skill_name, "unknown"


def scan_plugins() -> List[dict]:
    """Get approved and active skills from database with metadata from ZIP files.

    This is optimized to only show skills that are both approved and active,
    avoiding the need to scan all ZIP files repeatedly.
    """
    result = []

    # Get all approved and active skills directly from database
    from database import get_connection

    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                id, skill_name, version, filename, uploader_id, status,
                source_type, uploaded_at, reviewed_at, reviewer_id,
                review_comment, is_active, is_default_version
            FROM skills
            WHERE status = 'approved' AND is_active = 1
            ORDER BY skill_name, is_default_version DESC, version DESC
            """
        ).fetchall()

        # Deduplicate: only keep the first (default/latest) version for each skill
        seen_skills = set()
        deduplicated_rows = []
        for row in rows:
            if row["skill_name"] not in seen_skills:
                seen_skills.add(row["skill_name"])
                deduplicated_rows.append(row)
        rows = deduplicated_rows

        # Build result list with metadata from ZIP files
        for row in rows:
            skill_name = row["skill_name"]
            # Extract metadata from ZIP file
            metadata = extract_metadata(row["filename"])
            if not metadata:
                # Fallback if ZIP file doesn't exist or is invalid
                metadata = {
                    "name": skill_name,
                    "description": f"{skill_name} - 技能描述",
                    "version": row["version"],
                    "license": None,
                    "compatibility": None,
                    "metadata": {"version": row["version"], "author": "未知"},
                    "allowed_tools": None
                }

            # Get uploader employee_id from database
            uploader_id = row.get("uploader_id")
            uploader_employee_id = None
            if uploader_id:
                try:
                    uploader_row = conn.execute(
                        "SELECT employee_id FROM users WHERE id = %s",
                        (uploader_id,)
                    ).fetchone()
                    if uploader_row:
                        uploader_employee_id = uploader_row["employee_id"]
                except Exception:
                    pass

            # Use uploader employee_id as author if no author in metadata
            if uploader_employee_id:
                inner_meta = metadata.get("metadata", {})
                if not inner_meta.get("author"):
                    metadata["metadata"] = {**inner_meta, "author": uploader_employee_id}

            # Get file size
            file_size = 0
            plugins_dir = os.path.join(os.path.dirname(__file__), "plugins")
            file_path = os.path.join(plugins_dir, row["filename"])
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)

            # Extract category from metadata
            inner_meta = metadata.get("metadata", {})
            category = inner_meta.get("category") if isinstance(inner_meta, dict) else None

            result.append({
                "name": skill_name,
                "metadata": metadata,
                "category": category,  # Add category field
                "latest_version": row["version"],
                "source_type": row["source_type"] or "opensource",
                "uploaded_at": row["uploaded_at"].isoformat() if row["uploaded_at"] else None,
                "updated_at": row["uploaded_at"].strftime("%Y-%m-%d") if row["uploaded_at"] else None,
                "download_count": 0,  # TODO: Add download count if tracking
                "rating": 4.5,  # Default rating, TODO: Add actual rating
                "uploader_employee_id": uploader_employee_id,  # Add uploader info
                "versions": [{
                    "version": row["version"],
                    "filename": row["filename"],
                    "size": file_size,
                    "updated_at": row["uploaded_at"].strftime("%Y-%m-%d") if row["uploaded_at"] else None
                }]
            })

    return result


def extract_metadata(zip_filename: str) -> Optional[dict]:
    """Extract metadata from SKILL.md inside zip per Agent Skills specification.

    The ZIP should have structure:
        skill-name-1.0.0.zip
        └── skill-name/
            ├── SKILL.md
            ├── scripts/
            └── ...

    Args:
        zip_filename: Name of the ZIP file (e.g., "skill-name-1.0.0.zip")

    Returns:
        Metadata dict or fallback info
    """
    zip_path = PLUGINS_DIR / zip_filename
    skill_name, version = parse_plugin_filename(zip_filename)

    # 快速检查文件是否存在，避免不必要的 ZIP 操作和日志
    if not zip_path.exists():
        return {
            "name": skill_name,
            "description": f"{skill_name} - 技能描述",
            "version": version if version != "unknown" else "1.0.0",
            "author": None,
            "license": None,
            "compatibility": None,
            "metadata": {"version": version, "author": "未知"},
            "allowed_tools": None
        }

    # Try to extract from SKILL.md
    metadata = extract_metadata_from_skill_md(zip_path)

    if metadata:
        # Extract version from metadata field (per Agent Skills spec)
        skill_metadata = metadata.get("metadata", {})
        if isinstance(skill_metadata, dict):
            spec_version = skill_metadata.get("version")
        else:
            spec_version = None
            skill_metadata = {}

        # Normalize metadata format per Agent Skills spec
        # Author is in metadata.author from SKILL.md
        author = skill_metadata.get("author") or metadata.get("author")
        normalized = {
            "name": metadata.get("name", skill_name),
            "version": spec_version if spec_version else (version if version != "unknown" else "1.0.0"),
            "description": metadata.get("description", "No description available"),
            "author": author,  # Include author field
            "license": metadata.get("license"),
            "compatibility": metadata.get("compatibility"),
            "metadata": {**skill_metadata, "author": author},  # Include author in metadata for frontend
            "allowed_tools": metadata.get("allowed-tools")
        }
        return normalized

    # Fallback: try legacy package.json format for backward compatibility
    try:
        import zipfile
        with zipfile.ZipFile(zip_path, 'r') as zf:
            namelist = zf.namelist()

            for name in namelist:
                if name == 'package.json' or (name.endswith('/package.json')):
                    content = zf.read(name)
                    legacy_metadata = json.loads(content)
                    # Convert legacy format to Agent Skills format
                    legacy_author = legacy_metadata.get("author", {})
                    author_name = "Unknown"
                    if isinstance(legacy_author, dict):
                        author_name = legacy_author.get("name", "Unknown")
                    elif isinstance(legacy_author, str):
                        author_name = legacy_author

                    return {
                        "name": legacy_metadata.get("name", skill_name),
                        "version": legacy_metadata.get("version", version if version != "unknown" else "1.0.0"),
                        "description": legacy_metadata.get("description", "No description available"),
                        "license": None,
                        "compatibility": None,
                        "metadata": {
                            "author": author_name,
                            "version": legacy_metadata.get("version", version if version != "unknown" else "1.0.0")
                        },
                        "legacy": True
                    }
    except Exception:
        pass

    # Final fallback
    return {
        "name": skill_name,
        "version": version if version != "unknown" else "1.0.0",
        "description": "No description available",
        "license": None,
        "compatibility": None,
        "metadata": {"author": "未知"}
    }


# 隐藏安装指南页面
# @app.get("/install-guide", response_class=HTMLResponse)
# async def install_guide(request: Request):
#     """Installation guide page."""
#     return templates.TemplateResponse("install_guide.html", {
#         "request": request
#     })


# Markdown extensions for rendering docs
MARKDOWN_EXTENSIONS = ["tables", "fenced_code", "toc", "nl2br"]


@lru_cache(maxsize=1)
def _get_spec_html() -> str:
    """Load and render the spec markdown file once, cache result."""
    spec_file = Path(__file__).parent / "docs" / "skill_specification_v1.md"
    try:
        with open(spec_file, "r", encoding="utf-8") as f:
            return markdown.markdown(f.read(), extensions=MARKDOWN_EXTENSIONS)
    except (FileNotFoundError, OSError):
        return "<h1>规范文档未找到</h1><p>请联系管理员添加规范文档。</p>"



@app.get("/.well-known/skills/index.json")
async def skills_well_known_index(request: Request):
    """Skills index for npx skills CLI.

    Format follows vercel-labs/skills WellKnownIndex specification:
    - name: skill identifier (required)
    - description: skill description (required)
    - files: array of files in the skill (required, must include SKILL.md)
    """
    plugins = scan_plugins()

    skills_index = {
        "skills": []
    }

    for plugin in plugins:
        meta = plugin["metadata"]
        latest = plugin["versions"][-1]
        skill_name = meta.get("name", plugin["name"])

        # Normalize skill name: lowercase, alphanumeric and hyphens only
        normalized_name = skill_name.lower().replace("_", "-")
        # Remove any characters that aren't alphanumeric or hyphens
        import re
        normalized_name = re.sub(r'[^a-z0-9-]', '-', normalized_name)
        normalized_name = re.sub(r'-+', '-', normalized_name).strip('-')

        skills_index["skills"].append({
            "name": normalized_name,
            "description": meta.get("description", "No description"),
            "files": ["SKILL.md"]
        })

    return skills_index


@app.get("/.well-known/skills/{skill_name}/SKILL.md", response_class=PlainTextResponse)
async def get_skill_well_known(skill_name: str):
    """Serve SKILL.md file for npx skills CLI.

    Extracts SKILL.md from the skill ZIP file.
    """
    import zipfile
    import io

    plugins = scan_plugins()

    # Find the skill by normalized name
    for plugin in plugins:
        meta = plugin["metadata"]
        latest = plugin["versions"][-1]
        plugin_skill_name = meta.get("name", plugin["name"])

        # Normalize for comparison
        import re
        normalized = plugin_skill_name.lower().replace("_", "-")
        normalized = re.sub(r'[^a-z0-9-]', '-', normalized)
        normalized = re.sub(r'-+', '-', normalized).strip('-')

        if normalized == skill_name:
            # Found the skill, extract SKILL.md from ZIP
            zip_path = os.path.join(PLUGINS_DIR, latest["filename"])

            if not os.path.exists(zip_path):
                break

            try:
                with zipfile.ZipFile(zip_path, 'r') as zf:
                    # Look for SKILL.md (case-insensitive)
                    for name in zf.namelist():
                        if name.lower().endswith('skill.md'):
                            content = zf.read(name).decode('utf-8')
                            return content
            except Exception as e:
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail=f"Failed to read skill file: {str(e)}"
                )
            break

    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail=f"Skill '{skill_name}' not found"
    )


@app.get("/marketplace.json")
async def marketplace_json(request: Request):
    """Claude Code marketplace index."""
    plugins = scan_plugins()

    marketplace = {
        "name": "private-registry",
        "owner": {
            "name": "Internal Registry",
            "email": "admin@company.local"
        },
        "metadata": {
            "version": "1.0.0",
            "description": "Internal Claude Code Skill Registry",
            "updated_at": datetime.now().isoformat()
        },
        "plugins": []
    }

    for plugin in plugins:
        meta = plugin["metadata"]
        latest = plugin["versions"][-1]

        marketplace["plugins"].append({
            "name": meta.get("name", plugin["name"]),
            "version": latest["version"],
            "description": meta.get("description", "No description"),
            "author": meta.get("author", {"name": "Unknown"}),
            "download_url": f"http://{request.headers.get('host', 'localhost:28000')}/plugins/{latest['filename']}",
            "size_kb": round(latest["size"] / 1024, 1)
        })

    return marketplace


@app.get("/plugins/{filename}")
async def download_plugin(filename: str, request: Request):
    """Download plugin ZIP file (original uploaded file).

    Public endpoint - no authentication required for downloads.

    Args:
        filename: Name of the plugin file
        request: HTTP request

    Returns:
        Original ZIP file as uploaded by the user
    """
    # No authentication required - public download
    user_id = request.session.get("user_id")

    file_path = PLUGINS_DIR / filename

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Plugin not found")

    # Get skill name from filename (version is now in SKILL.md metadata)
    skill_name = filename[:-4] if filename.endswith('.zip') else filename

    # Extract version from SKILL.md inside the ZIP for download tracking
    metadata = extract_metadata_from_skill_md(file_path)
    if metadata and metadata.get("metadata"):
        version = metadata.get("metadata", {}).get("version", "unknown")
    else:
        version = "unknown"

    # Record download
    try:
        record_download(
            skill_name=skill_name,
            version=version,
            filename=filename,
            ip_address=request.client.host if request.client else None,
            user_agent=request.headers.get("user-agent"),
            user_id=user_id
        )
    except Exception as e:
        # Log error but don't block download
        logger.warning(f"Failed to record download: {e}", extra={"skill_name": skill_name, "filename": filename})

    # Return original ZIP file
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/zip"
    )


    """User login API endpoint.

    Accepts employee_id and api_key as form parameters.
    Sets session variables on success.
    Updates last login timestamp.
    Returns success response or redirects on failure.
    """
    # Query user by credentials
    user = get_user_by_credentials(employee_id, api_key)

    if user:
        # Set session variables
        request.session["user_id"] = user["id"]
        request.session["employee_id"] = user["employee_id"]
        request.session["role"] = user["role"]

        # Update last login
        update_last_login(user["id"])

        # Redirect to homepage after login
        return RedirectResponse(url="/", status_code=302)
    else:
        return RedirectResponse(
            url="/login?error=invalid",
            status_code=302
        )


@app.get("/logout")
async def logout(request: Request):
    """Logout and clear session."""
    request.session.clear()
    return RedirectResponse(url="/", status_code=302)


@app.get("/api/me")
async def api_me(request: Request):
    """Get current user information.

    Returns the current authenticated user's details.
    Raises HTTP 401 if not authenticated.
    """
    user = get_current_user(request)

    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Not authenticated"
        )

    return {
        "id": user["id"],
        "employee_id": user["employee_id"],
        "role": user["role"],
        "created_at": user["created_at"],
        "last_login": user["last_login"]
    }


@app.get("/upload", response_class=HTMLResponse)
async def user_upload_page(request: Request):
    """Display user upload page (requires auth)."""
    # Get current user
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    return templates.TemplateResponse("upload.html", {
        "request": request,
        "user": user,
        "success": None,
        "error": None
    })


@app.get("/admin/upload", response_class=HTMLResponse)
async def upload_page(request: Request, _: bool = Depends(require_auth)):
    """Display admin upload page (requires auth)."""
    user = get_current_user(request)

    return templates.TemplateResponse("admin_upload.html", {
        "request": request,
        "user": user,
        "success": None,
        "error": None
    })


@app.get("/admin", response_class=HTMLResponse)
async def admin_dashboard(request: Request):
    """Display admin dashboard (requires admin)."""
    # Get current user
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    # Check admin role
    if user["role"] != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )

    return templates.TemplateResponse("admin.html", {
        "request": request,
        "user": user
    })


@app.get("/admin/users", response_class=HTMLResponse)
async def admin_users_page(request: Request):
    """Display user management page (requires admin)."""
    # Get current user
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    # Check admin role
    if user["role"] != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )

    return templates.TemplateResponse("admin_users.html", {
        "request": request,
        "user": user
    })


@app.get("/admin/api-keys", response_class=HTMLResponse)
async def admin_api_keys_page(request: Request):
    """Display API Keys management page (requires admin)."""
    # Get current user
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    # Check admin role
    if user["role"] != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )

    return templates.TemplateResponse("admin_api_keys.html", {
        "request": request,
        "user": user
    })


def validate_skill_zip(zip_path: Path, allow_missing: bool = False, default_author: str = None) -> tuple[bool, dict]:
    """Validate a skill ZIP file according to Agent Skills specification.

    Args:
        zip_path: Path to the ZIP file
        allow_missing: If True, return missing fields info instead of rejecting
        default_author: Default author to use if not specified in SKILL.md

    The ZIP should have structure:
        skill-name/
        ├── SKILL.md          # Required
        ├── scripts/          # Optional
        ├── references/       # Optional
        └── assets/           # Optional

    SKILL.md must contain YAML frontmatter with required fields:
        - name: skill identifier (max 64 chars, lowercase letters/numbers/hyphens only)
        - description: what the skill does (max 1024 chars)
        - metadata.version: version string (e.g., "1.0.0")
        - metadata.author: author identifier (format: lowercase letter + 8 digits, e.g., "w00545471")

    Optional fields:
        - license: license name or reference
        - compatibility: environment requirements (max 500 chars)
        - metadata: arbitrary key-value mapping (other custom fields)
        - allowed-tools: space-delimited list of pre-approved tools

    Args:
        zip_path: Path to the skill ZIP file
        allow_missing: If True, return missing fields info instead of rejecting

    Returns:
        (is_valid, metadata or error_info)
        When allow_missing=True and fields are missing, returns:
        (False, {"error": "MISSING_FIELDS", "missing_fields": [...], "metadata": {...}})
    """
    import zipfile

    try:
        with zipfile.ZipFile(zip_path, 'r') as zf:
            namelist = zf.namelist()

            # Find SKILL.md (may be in root or subdirectory)
            skill_md_paths = [name for name in namelist
                             if name.endswith('SKILL.md') or name == 'SKILL.md']

            if not skill_md_paths:
                return False, {"error": "Missing SKILL.md in ZIP"}

            # Read and parse SKILL.md
            skill_md_path = skill_md_paths[0]
            content = zf.read(skill_md_path).decode('utf-8')
            metadata, markdown_body = parse_skill_md(content)

            warnings = []

            # If YAML parsing fails, try to extract info from content
            if metadata is None:
                warnings.append("SKILL.md 的 YAML 格式无法解析，将使用默认值")
                metadata = {}

            # Extract name - try to get from ZIP filename if missing
            skill_name = metadata.get("name")
            if not skill_name:
                # Try to extract from ZIP filename (e.g., skill-name-1.0.0.zip -> skill-name)
                zip_name = zip_path.stem
                import re
                # Remove version suffix like -1.0.0, -1.0, etc.
                skill_name = re.sub(r'-\d+(\.\d+)*$', '', zip_name)
                warnings.append(f"未找到 name 字段，将使用文件名: {skill_name}")

            # Extract description - use markdown body or default
            description = metadata.get("description")
            if not description:
                # Use first line of markdown body as description
                if markdown_body:
                    first_line = markdown_body.strip().split('\n')[0][:200]
                    description = first_line if first_line else f"Skill: {skill_name}"
                else:
                    description = f"Skill: {skill_name}"
                warnings.append("未找到 description 字段，将使用默认描述")

            # Validate name format - DISABLED: 取消name格式校验
            # Validate description length (max 1024 chars)
            if not isinstance(description, str) or len(description) == 0 or len(description) > 1024:
                description = description[:1024] if description else f"Skill: {skill_name}"

            # Validate optional fields if present
            # compatibility: max 500 chars
            if "compatibility" in metadata:
                compat = metadata["compatibility"]
                if not isinstance(compat, str) or len(compat) == 0 or len(compat) > 500:
                    return False, {"error": "Compatibility must be 1-500 characters if provided"}

            # Extract metadata fields (version and author are optional, will use defaults if missing)
            skill_metadata = metadata.get("metadata") or {}
            # If metadata is None or not a dict (e.g., empty in YAML), use empty dict
            if skill_metadata and not isinstance(skill_metadata, dict):
                return False, {"error": "Metadata must be a key-value mapping"}

            # 获取 version 和 author，如果未填写则使用默认值
            version = skill_metadata.get("version") or "1.0.0"  # 默认版本号
            # author 如果未填写，使用传入的 default_author（上传用户的id）
            author = skill_metadata.get("author") or default_author

            # Normalize metadata for return (matching API format)
            normalized_metadata = {
                "name": skill_name,
                "description": description,
                "version": version,
                "author": author,
                "license": metadata.get("license"),
                "compatibility": metadata.get("compatibility"),
                "metadata": skill_metadata,
                "allowed_tools": metadata.get("allowed-tools")
            }

            # Add warnings to metadata if any
            if warnings:
                normalized_metadata["_warnings"] = warnings

            return True, normalized_metadata

    except zipfile.BadZipFile:
        return False, {"error": "Invalid ZIP file"}
    except yaml.YAMLError as e:
        return False, {"error": f"Invalid YAML in SKILL.md: {str(e)}"}
    except Exception as e:
        return False, {"error": str(e)}


def save_skill_zip(temp_zip: Path, metadata: dict) -> Path:
    """Save a skill ZIP to the plugins directory.

    Args:
        temp_zip: Path to the temporary ZIP file
        metadata: Skill metadata from SKILL.md

    Returns:
        Path to the saved ZIP file
    """
    skill_name = metadata["name"]
    version = metadata.get("version", "1.0.0")
    target_filename = f"{skill_name}-{version}.zip"
    target_path = PLUGINS_DIR / target_filename

    # Copy file to target location
    shutil.copy(temp_zip, target_path)

    return target_path


def approve_skill_file(skill_id: int) -> bool:
    """Approve a skill by moving it from pending to plugins directory.

    In single-version mode, if an approved version of the same skill_name exists,
    it will be removed before the new version is moved.

    Args:
        skill_id: The ID of the skill to approve

    Returns:
        True if successful, False otherwise
    """
    # Get skill record
    skill = get_skill_by_id(skill_id)
    if not skill:
        return False

    if skill["status"] != "pending":
        return False

    # File paths
    pending_path = PENDING_DIR / skill["filename"]
    plugins_path = PLUGINS_DIR / skill["filename"]
    skill_name = skill["skill_name"]

    try:
        # In single-version mode, check for existing approved version of this skill
        from database import get_skill_by_name
        existing_skill = get_skill_by_name(skill_name)

        # If there's an existing approved version, delete its file and record
        if existing_skill and existing_skill["status"] == "approved" and existing_skill["id"] != skill_id:
            old_file_path = PLUGINS_DIR / existing_skill["filename"]
            if old_file_path.exists():
                logger.info(f"Removing old version file: {old_file_path}")
                old_file_path.unlink()
            # Delete the old database record
            from database import get_connection
            with get_connection() as conn:
                conn.execute("DELETE FROM skills WHERE id = %s", (existing_skill["id"],))
                conn.commit()
            logger.info(f"Deleted old skill record: {existing_skill['id']}")

        # Check if file is in pending directory
        if pending_path.exists():
            # Move file from pending to plugins
            # Remove existing file if it exists (prevents FileExistsError on re-approval)
            if plugins_path.exists():
                logger.info(f"Removing existing file: {plugins_path}")
                plugins_path.unlink()

            shutil.move(str(pending_path), str(plugins_path))
        elif not plugins_path.exists():
            # File not in pending and not in plugins - error
            logger.error(f"Skill file not found: {skill['filename']} (checked both pending and plugins directories)")
            return False
        else:
            # File already in plugins directory (possibly from old batch upload)
            logger.info(f"File already in plugins directory: {plugins_path}")

        # Update database status
        update_skill_status(skill_id, "approved")

        # Set is_active=1 on approval
        update_skill_active_status(skill_id, True)

        # Create notification for uploader
        uploader_id = skill.get("uploader_id")
        if uploader_id:
            content = f"您的技能 {skill['skill_name']} (版本 {skill['version']}) 已通过审核并上线。"
            create_notification(
                user_id=uploader_id,
                type="review_success",
                title="您的技能已通过审核",
                content=content,
                related_skill_id=skill_id
            )

        return True
    except Exception as e:
        logger.error(f"Failed to approve skill {skill_id}: {e}")
        return False


def reject_skill_file(skill_id: int, comment: Optional[str] = None) -> bool:
    """Reject a skill by deleting the pending file.

    Args:
        skill_id: The ID of the skill to reject
        comment: Optional rejection comment

    Returns:
        True if successful, False otherwise
    """
    # Get skill record
    skill = get_skill_by_id(skill_id)
    if not skill:
        return False

    if skill["status"] != "pending":
        return False

    # Delete pending file
    pending_path = PENDING_DIR / skill["filename"]

    try:
        if pending_path.exists():
            pending_path.unlink()

        # Update database status
        update_skill_status(skill_id, "rejected", comment=comment)

        # Create notification for uploader
        uploader_id = skill.get("uploader_id")
        if uploader_id:
            content = f"您的技能 {skill['skill_name']} (版本 {skill['version']}) 未通过审核。"
            if comment:
                content += f" 原因: {comment}"
            create_notification(
                user_id=uploader_id,
                type="review_rejected",
                title="您的技能未通过审核",
                content=content,
                related_skill_id=skill_id
            )

        return True
    except Exception as e:
        logger.error(f"Failed to reject skill {skill_id}: {e}")
        return False


@app.get("/api/pending")
async def api_pending_skills(
    _: bool = Depends(require_admin)
):
    """Get all pending skills awaiting approval (admin only)."""
    try:
        pending = get_pending_skills()
        return {
            "success": True,
            "data": pending,
            "count": len(pending)
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch pending skills: {str(e)}"
        )


@app.post("/api/review/{skill_id}")
async def api_review_skill(
    skill_id: int,
    request: Request,
    _: bool = Depends(require_admin)
):
    """Approve or reject a pending skill (admin only).

    Expects JSON body with:
    {
        "action": "approve" | "reject",
        "comment": "optional comment"
    }
    """
    try:
        # Get current user (reviewer)
        reviewer_id = request.session.get("user_id")
        if not reviewer_id:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Not authenticated"
            )

        # Parse request body
        data = await request.json()
        action = data.get("action")
        comment = data.get("comment")

        if action not in ["approve", "reject"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid action. Must be 'approve' or 'reject'"
            )

        # Get skill record
        skill = get_skill_by_id(skill_id)
        if not skill:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Skill {skill_id} not found"
            )

        if skill["status"] != "pending":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Skill {skill_id} is not in pending status"
            )

        # Perform action
        if action == "approve":
            # Re-check status before approval to handle race conditions
            skill = get_skill_by_id(skill_id)
            if not skill or skill["status"] != "pending":
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Skill {skill_id} is no longer pending, may have been processed"
                )

            # Proceed with approval
            success = approve_skill_file(skill_id)
            if not success:
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail=f"Failed to approve skill {skill_id}"
                )

            # Update with reviewer info
            update_skill_status(skill_id, "approved", reviewer_id=reviewer_id, comment=comment)

            # NEW: Create Gitea push task
            task_id = None
            try:
                from services.gitea.gitea_integration import create_push_task
                task_id = create_push_task(skill_id)
                logger.info(f"Created Gitea push task {task_id} for skill {skill_id}")
            except Exception as e:
                # Log error but don't block approval
                logger.error(f"Failed to create Gitea push task: {e}")

            return {
                "success": True,
                "message": f"Skill {skill['skill_name']}@{skill['version']} approved",
                "skill_id": skill_id,
                "push_task_id": task_id
            }

        else:  # reject
            success = reject_skill_file(skill_id, comment)
            if not success:
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail=f"Failed to reject skill {skill_id}"
                )

            # Update with reviewer info
            update_skill_status(skill_id, "rejected", reviewer_id=reviewer_id, comment=comment)

            return {
                "success": True,
                "message": f"Skill {skill['skill_name']}@{skill['version']} rejected",
                "skill_id": skill_id
            }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Review failed: {str(e)}"
        )


@app.get("/api/user/downloads")
async def api_user_downloads(
    request: Request,
    page: int = Query(1, ge=1, description="Page number (starts from 1)"),
    per_page: int = Query(20, ge=1, le=100, description="Items per page (max 100)"),
    _: bool = Depends(require_auth)
):
    """Get the current user's download history with pagination.

    Query parameters:
    - page: Page number (default: 1, min: 1)
    - per_page: Items per page (default: 20, min: 1, max: 100)

    Returns paginated list of downloads for the authenticated user.
    """
    try:
        # Get current user
        user_id = request.session.get("user_id")
        if not user_id:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Not authenticated"
            )

        # Calculate offset from page number
        offset = (page - 1) * per_page

        # Get user downloads from database
        result = get_user_downloads(
            user_id=user_id,
            limit=per_page,
            offset=offset
        )

        # Calculate pagination metadata
        total = result["total"]
        total_pages = (total + per_page - 1) // per_page

        return {
            "success": True,
            "data": result["downloads"],
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total": total,
                "total_pages": total_pages
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch downloads: {str(e)}"
        )


@app.get("/api/user/uploads")
async def api_user_uploads(
    request: Request,
    _: bool = Depends(require_auth)
):
    """Get the current user's upload history."""
    try:
        user_id = request.session.get("user_id")
        if not user_id:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Not authenticated"
            )

        uploads = get_user_uploads(user_id)

        return {
            "success": True,
            "data": uploads,
            "count": len(uploads)
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch uploads: {str(e)}"
        )


@app.get("/api/notifications")
async def api_get_notifications(
    request: Request,
    unread_only: bool = Query(False, description="Filter to only unread notifications"),
    limit: int = Query(50, ge=1, le=100, description="Maximum number of notifications to return"),
    offset: int = Query(0, ge=0, description="Number of notifications to skip"),
    _: bool = Depends(require_auth)
):
    """Get notifications for the current user with pagination.

    Query parameters:
    - unread_only: If True, only return unread notifications (default: False)
    - limit: Maximum number of notifications to return (default: 50, max: 100)
    - offset: Number of notifications to skip for pagination (default: 0)

    Returns notifications sorted newest first.
    """
    try:
        user_id = request.session.get("user_id")
        if not user_id:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Not authenticated"
            )

        result = get_user_notifications(user_id, unread_only, limit, offset)

        return {
            "success": True,
            "data": result
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch notifications: {str(e)}"
        )


@app.get("/api/notifications/unread-count")
async def api_get_unread_count(
    request: Request,
    _: bool = Depends(require_auth)
):
    """Get the count of unread notifications for the current user."""
    try:
        user_id = request.session.get("user_id")
        if not user_id:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Not authenticated"
            )

        count = get_unread_notifications_count(user_id)

        return {
            "success": True,
            "data": {
                "unread_count": count
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch unread count: {str(e)}"
        )


@app.post("/api/notifications/{notification_id}/read")
async def api_mark_notification_read(
    notification_id: int,
    request: Request,
    _: bool = Depends(require_auth)
):
    """Mark a specific notification as read.

    Verifies that the user owns this notification before marking it as read.
    """
    try:
        user_id = request.session.get("user_id")
        if not user_id:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Not authenticated"
            )

        success = mark_notification_read(notification_id, user_id)

        if not success:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Notification not found or does not belong to this user"
            )

        return {
            "success": True,
            "data": {
                "message": "Notification marked as read"
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to mark notification as read: {str(e)}"
        )


@app.post("/api/notifications/read-all")
async def api_mark_all_read(
    request: Request,
    _: bool = Depends(require_auth)
):
    """Mark all notifications as read for the current user."""
    try:
        user_id = request.session.get("user_id")
        if not user_id:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Not authenticated"
            )

        count = mark_all_notifications_read(user_id)

        return {
            "success": True,
            "data": {
                "message": f"Marked {count} notifications as read",
                "count": count
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to mark all notifications as read: {str(e)}"
        )


@app.get("/api/my-skills")
async def api_my_skills(
    request: Request,
    status: str = Query("all", description="Filter by status: all, active, unlisted, pending, rejected"),
    page: int = Query(1, ge=1, description="Page number (starts from 1)"),
    per_page: int = Query(20, ge=1, le=100, description="Items per page (max 100)"),
    _: bool = Depends(require_auth)
):
    """Get current user's skills with pagination and filtering.

    Query parameters:
    - status: Filter by status ('all', 'active', 'unlisted', 'pending', 'rejected')
    - page: Page number (default: 1, min: 1)
    - per_page: Items per page (default: 20, min: 1, max: 100)

    Returns paginated list of skills for the authenticated user.
    """
    try:
        user_id = request.session.get("user_id")
        if not user_id:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Not authenticated"
            )

        # Calculate offset from page number
        offset = (page - 1) * per_page

        # Get user skills from database
        result = get_my_skills(
            user_id=user_id,
            status_filter=status,
            limit=per_page,
            offset=offset
        )

        # Calculate pagination metadata
        total = result["total"]
        total_pages = (total + per_page - 1) // per_page

        return {
            "success": True,
            "data": result["skills"],
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total": total,
                "total_pages": total_pages
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch skills: {str(e)}"
        )


@app.get("/my-skills", response_class=HTMLResponse)
async def my_skills_page(request: Request):
    """Render the my_skills.html page (requires auth)."""
    user = get_current_user(request)

    return templates.TemplateResponse("my_skills.html", {
        "request": request,
        "user": user
    })


class BatchOperationRequest(BaseModel):
    """Request model for batch operations."""
    skill_ids: List[int]


@app.post("/api/my-skills/batch/unlist")
async def api_batch_unlist_skills(
    request_data: BatchOperationRequest,
    request: Request,
    _: bool = Depends(require_auth)
):
    """Unlist multiple skills at once.

    User must own all the skills to unlist them.
    """
    try:
        user_id = request.session.get("user_id")
        if not user_id:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Not authenticated"
            )

        result = batch_unlist_skills(user_id, request_data.skill_ids)

        message = f"已下架 {result['success_count']} 个技能"
        if result["failed_ids"]:
            message += f"，{len(result['failed_ids'])} 个失败"

        return {
            "success": True,
            "message": message,
            "success_count": result["success_count"],
            "failed_ids": result["failed_ids"]
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to batch unlist skills: {str(e)}"
        )


@app.post("/api/my-skills/batch/delete")
async def api_batch_delete_skills(
    request_data: BatchOperationRequest,
    request: Request,
    _: bool = Depends(require_admin)
):
    """Delete multiple skills at once (admin only).

    Only admin users can delete any skills. The physical ZIP files will also be removed.
    """
    try:
        user_id = request.session.get("user_id")
        if not user_id:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Not authenticated"
            )

        result = batch_delete_skills(user_id, request_data.skill_ids)

        message = f"已删除 {result['success_count']} 个技能"
        if result["failed_ids"]:
            message += f"，{len(result['failed_ids'])} 个失败"

        return {
            "success": True,
            "message": message,
            "success_count": result["success_count"],
            "failed_ids": result["failed_ids"]
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to batch delete skills: {str(e)}"
        )


@app.post("/api/my-skills/{skill_id}/unlist")
async def api_unlist_skill(
    skill_id: int,
    request: Request,
    _: bool = Depends(require_auth)
):
    """Unlist a skill (set is_active = 0).

    User must own the skill to unlist it.
    """
    try:
        user_id = request.session.get("user_id")
        if not user_id:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Not authenticated"
            )

        # Get skill record
        skill = get_skill_by_id(skill_id)
        if not skill:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Skill {skill_id} not found"
            )

        # Verify ownership
        if skill["uploader_id"] != user_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You don't own this skill"
            )

        # Update active status
        update_skill_active_status(skill_id, False)

        return {
            "success": True,
            "message": f"Skill {skill['skill_name']}@{skill['version']} has been unlisted",
            "skill_id": skill_id
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to unlist skill: {str(e)}"
        )


@app.post("/api/my-skills/{skill_id}/publish")
async def api_publish_skill(
    skill_id: int,
    request: Request,
    _: bool = Depends(require_auth)
):
    """Publish a skill (set is_active = 1).

    User must own the skill to publish it.
    """
    try:
        user_id = request.session.get("user_id")
        if not user_id:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Not authenticated"
            )

        # Get skill record
        skill = get_skill_by_id(skill_id)
        if not skill:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Skill {skill_id} not found"
            )

        # Verify ownership
        if skill["uploader_id"] != user_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You don't own this skill"
            )

        # Update active status
        update_skill_active_status(skill_id, True)

        return {
            "success": True,
            "message": f"Skill {skill['skill_name']}@{skill['version']} has been published",
            "skill_id": skill_id
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to publish skill: {str(e)}"
        )


@app.delete("/api/my-skills/{skill_id}")
async def api_delete_skill(
    skill_id: int,
    request: Request,
    _: bool = Depends(require_admin)
):
    """Delete a skill version (admin only).

    Only admin users can delete any skill. The physical ZIP file will also be removed.
    If this is the default version and there are other versions, another version will be set as default.
    """
    try:
        user_id = request.session.get("user_id")
        if not user_id:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Not authenticated"
            )

        # Get skill record for response message
        skill = get_skill_by_id(skill_id)
        if not skill:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Skill {skill_id} not found"
            )

        # Admin can delete any skill, no ownership check needed
        # Delete the skill (pass is_admin=True to skip ownership check)
        success = delete_skill_version(user_id, skill_id, is_admin=True)

        if not success:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to delete skill"
            )

        return {
            "success": True,
            "message": f"Skill {skill['skill_name']}@{skill['version']} has been deleted",
            "skill_id": skill_id
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete skill: {str(e)}"
        )


@app.post("/api/upload")
async def upload_plugin(
    request: Request,
    file: UploadFile = File(...),
    source_type: str = Form(default="opensource"),
    overwrite: bool = Form(default=False),
    _: bool = Depends(require_auth)
):
    """Upload a single skill ZIP file (requires auth).

    Saves to pending directory and creates database record with status='pending'.
    Requires admin approval before being made available.

    Returns JSON response for AJAX requests.
    """
    import tempfile
    from fastapi.responses import JSONResponse

    # Get current user
    user_id = request.session.get("user_id")
    if not user_id:
        return JSONResponse(
            status_code=status.HTTP_401_UNAUTHORIZED,
            content={"success": False, "error": "请先登录"}
        )

    # Validate file extension
    if not file.filename or not file.filename.endswith('.zip'):
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "error": "只支持 ZIP 格式的文件"}
        )

    # Save uploaded file to temp location
    temp_dir = tempfile.mkdtemp()
    temp_zip = Path(temp_dir) / "upload.zip"

    try:
        with open(temp_zip, "wb") as f:
            shutil.copyfileobj(file.file, f)

        # Validate the ZIP file (allow missing fields, use current user as default author)
        is_valid, result = validate_skill_zip(temp_zip, allow_missing=True, default_author=user_id)

        if not is_valid:
            error_msg = result.get('error', 'Unknown error')

            # 处理缺失字段的情况
            if error_msg == "MISSING_FIELDS":
                # 生成临时文件 ID 并保存文件
                import secrets
                temp_file_id = secrets.token_urlsafe(16)
                temp_storage_dir = Path(tempfile.gettempdir()) / "skillhub_uploads"
                temp_storage_dir.mkdir(exist_ok=True)
                temp_storage_path = temp_storage_dir / f"{temp_file_id}.zip"
                shutil.copy(temp_zip, temp_storage_path)

                # 返回缺失字段信息
                return JSONResponse(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    content={
                        "success": False,
                        "error": "MISSING_FIELDS",
                        "data": {
                            "missing_fields": result.get("missing_fields", []),
                            "metadata": result.get("metadata", {}),
                            "temp_file_id": temp_file_id
                        }
                    }
                )

            # Map validation errors to user-friendly messages
            error_messages = {
                "Missing SKILL.md in ZIP": "ZIP 文件中缺少 SKILL.md 文件",
                "Invalid YAML frontmatter in SKILL.md": "SKILL.md 中的 YAML 格式无效",
                "Missing required field 'name' in SKILL.md YAML frontmatter": "SKILL.md 中缺少必需字段 'name'",
                "Missing required field 'description' in SKILL.md YAML frontmatter": "SKILL.md 中缺少必需字段 'description'",
                "Missing required field 'metadata.version' in SKILL.md": "SKILL.md 中缺少 metadata.version 字段",
                "Metadata.version must be a non-empty string": "版本号不能为空",
                "Missing required field 'metadata.author' in SKILL.md": "SKILL.md 中缺少 metadata.author 字段",
                "Metadata.author must be a non-empty string": "作者信息不能为空",
                "Invalid ZIP file": "无效的 ZIP 文件",
            }

            # Check for skill name validation errors
            if "Invalid skill name:" in error_msg:
                return JSONResponse(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    content={"success": False, "error": f"技能名称格式错误: {error_msg.split(':', 1)[1].strip() if ':' in error_msg else error_msg}"}
                )

            # Use user-friendly message if available, otherwise use original
            user_friendly_error = error_messages.get(error_msg, error_msg)

            # Return JSON error for all cases
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content={"success": False, "error": user_friendly_error}
            )

        # Save the skill ZIP to pending directory
        skill_name = result["name"]
        version = result.get("version", "1.0.0")
        target_filename = f"{skill_name}-{version}.zip"
        target_path = PENDING_DIR / target_filename

        # Check if skill with same name AND version already exists
        from database import get_connection
        with get_connection() as conn:
            existing_skill = conn.execute("""
                SELECT id, skill_name, version, uploader_id, status
                FROM skills
                WHERE skill_name = %s AND version = %s
                LIMIT 1
            """, (skill_name, version)).fetchone()

        if existing_skill:
            # Same name and version exists
            if existing_skill["uploader_id"] != user_id:
                error_msg = f"技能 {skill_name} 已被其他用户创建，您无权更新"
                return JSONResponse(
                    status_code=status.HTTP_403_FORBIDDEN,
                    content={"success": False, "error": error_msg}
                )
            # Skill exists with same version - ask for confirmation to overwrite
            if not overwrite:
                return JSONResponse(
                    status_code=status.HTTP_409_CONFLICT,
                    content={
                        "success": False,
                        "error": "SKILL_EXISTS",
                        "message": f"技能 {skill_name} v{version} 已存在，是否覆盖更新？",
                        "skill_name": skill_name,
                        "existing_version": version
                    }
                )
            # Delete the old pending file if exists
            old_filename = f"{skill_name}-{version}.zip"
            old_pending_path = PENDING_DIR / old_filename
            if old_pending_path.exists():
                old_pending_path.unlink()

            # Update existing record with new version and file
            from database import get_connection
            with get_connection() as conn:
                conn.execute(
                    """
                    UPDATE skills
                    SET version = %s, filename = %s, status = 'pending',
                        uploader_id = %s, source_type = %s
                    WHERE id = %s
                    """,
                    (version, target_filename, user_id, source_type, existing_skill['id'])
                )
                conn.commit()
            skill_id = existing_skill['id']

            # Copy file to pending location
            shutil.copy(temp_zip, target_path)

            # Return success response
            success_msg = f"成功更新 {result['name']}@{result['version']}，等待管理员审核"

            # Check for warnings
            upload_warnings = result.get("_warnings", [])

            # Return JSON for all cases
            response_content = {
                "success": True,
                "message": success_msg,
                "skill_name": result['name'],
                "version": result['version'],
                "skill_id": skill_id
            }
            if upload_warnings:
                response_content["warnings"] = upload_warnings

            return JSONResponse(content=response_content)

        # Copy file to pending location
        shutil.copy(temp_zip, target_path)

        # Create database record with status='pending'
        from database import create_skill_record
        skill_id = create_skill_record(
            skill_name=skill_name,
            version=version,
            filename=target_filename,
            uploader_id=user_id,
            status='pending',
            source_type=source_type
        )

        # Return success response
        success_msg = f"成功上传 {result['name']}@{result['version']}，等待管理员审核"

        # Check for warnings
        upload_warnings = result.get("_warnings", [])

        # Return JSON for all cases
        response_content = {
            "success": True,
            "message": success_msg,
            "skill_name": result['name'],
            "version": result['version'],
            "skill_id": skill_id
        }
        if upload_warnings:
            response_content["warnings"] = upload_warnings

        return JSONResponse(content=response_content)

    except Exception as e:
        import traceback
        traceback.print_exc()
        error_msg = f"上传失败: {str(e)}"

        if "admin" in request.headers.get("referer", ""):
            return templates.TemplateResponse("admin_upload.html", {
                "request": request,
                "success": None,
                "error": error_msg
            })
        else:
            return JSONResponse(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                content={"success": False, "error": error_msg}
            )

    finally:
        # Cleanup
        shutil.rmtree(temp_dir, ignore_errors=True)


def update_skill_metadata_in_zip(zip_path: Path, version: str, author: str) -> Path:
    """更新 ZIP 文件中的 SKILL.md，添加缺失的 metadata.version 和 metadata.author

    Args:
        zip_path: 原始 ZIP 文件路径
        version: 版本号
        author: 作者ID

    Returns:
        更新后的 ZIP 文件路径
    """
    import zipfile
    import tempfile

    try:
        # 读取原始 ZIP
        with zipfile.ZipFile(zip_path, 'r') as zf:
            namelist = zf.namelist()

            # 找到 SKILL.md
            skill_md_paths = [name for name in namelist
                             if name.endswith('SKILL.md') or name == 'SKILL.md']

            if not skill_md_paths:
                raise ValueError("SKILL.md not found in ZIP")

            skill_md_path = skill_md_paths[0]
            original_content = zf.read(skill_md_path).decode('utf-8')

        # 解析并更新 YAML frontmatter
        metadata, content = parse_skill_md(original_content)

        if metadata is None:
            raise ValueError("Invalid YAML frontmatter in SKILL.md")

        # 确保 metadata 字典存在
        if 'metadata' not in metadata or not isinstance(metadata['metadata'], dict):
            metadata['metadata'] = {}

        # 更新 version 和 author
        metadata['metadata']['version'] = version
        metadata['metadata']['author'] = author

        # 重建 SKILL.md 内容
        import yaml
        yaml_frontmatter = yaml.dump(metadata, default_flow_style=False, allow_unicode=True)
        updated_content = f"---\n{yaml_frontmatter}---\n\n{content}"

        # 创建新的 ZIP 文件
        output_path = zip_path.parent / f"updated_{zip_path.name}"

        with zipfile.ZipFile(zip_path, 'r') as zf_in:
            with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zf_out:
                for item in zf_in.infolist():
                    if item.filename == skill_md_path:
                        # 写入更新后的 SKILL.md
                        zf_out.writestr(item, updated_content.encode('utf-8'))
                    else:
                        # 复制其他文件
                        zf_out.writestr(item, zf_in.read(item.filename))

        # 替换原文件
        shutil.move(output_path, zip_path)

        return zip_path

    except Exception as e:
        raise ValueError(f"Failed to update ZIP metadata: {str(e)}")


@app.post("/api/upload/complete")
async def complete_upload_with_metadata(
    request: Request,
    temp_file_id: str = Form(...),
    version: str = Form(...),
    author: str = Form(...),
    source_type: str = Form(default="opensource"),
    _: bool = Depends(require_auth)
):
    """补充缺失的 metadata 信息后完成上传

    Args:
        temp_file_id: 临时文件 ID
        version: 版本号
        author: 作者ID
        source_type: 来源类型

    Returns:
        上传结果
    """
    import tempfile
    from fastapi.responses import JSONResponse

    # 获取当前用户
    user_id = request.session.get("user_id")
    if not user_id:
        return JSONResponse(
            status_code=status.HTTP_401_UNAUTHORIZED,
            content={"success": False, "error": "请先登录"}
        )

    # 验证作者格式
    import re
    if not re.match(r'^[a-z]\d{8}$', author):
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "error": "作者ID格式错误，必须是小写字母+8位数字（如：w00545471）"}
        )

    # 验证版本号格式
    if not version or not isinstance(version, str) or len(version) == 0:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "error": "版本号不能为空"}
        )

    # 获取临时文件路径
    temp_storage_dir = Path(tempfile.gettempdir()) / "skillhub_uploads"
    temp_zip = temp_storage_dir / f"{temp_file_id}.zip"

    if not temp_zip.exists():
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"success": False, "error": "临时文件不存在，请重新上传"}
        )

    try:
        # 更新 ZIP 文件中的 metadata
        update_skill_metadata_in_zip(temp_zip, version, author)

        # 重新验证 ZIP 文件
        is_valid, result = validate_skill_zip(temp_zip, allow_missing=False, default_author=user_id)

        if not is_valid:
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content={"success": False, "error": result.get('error', '验证失败')}
            )

        # 保存到 pending 目录
        skill_name = result["name"]
        target_filename = f"{skill_name}-{version}.zip"
        target_path = PENDING_DIR / target_filename

        # 检查是否已存在
        from database import get_skill_by_name
        existing_skill = get_skill_by_name(skill_name)
        if existing_skill:
            if existing_skill["uploader_id"] != user_id:
                return JSONResponse(
                    status_code=status.HTTP_403_FORBIDDEN,
                    content={"success": False, "error": f"技能 {skill_name} 已被其他用户创建，您无权更新"}
                )
            # 对于已存在的技能，这是更新操作，继续执行

        # 复制文件到 pending 位置
        shutil.copy(temp_zip, target_path)

        # 创建数据库记录
        from database import create_skill_record
        skill_id = create_skill_record(
            skill_name=skill_name,
            version=version,
            filename=target_filename,
            uploader_id=user_id,
            status='pending',
            source_type=source_type
        )

        return JSONResponse(
            content={
                "success": True,
                "message": f"成功上传 {skill_name}@{version}，等待管理员审核",
                "skill_name": skill_name,
                "version": version,
                "skill_id": skill_id
            }
        )

    except ValueError as e:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "error": str(e)}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "error": f"上传失败: {str(e)}"}
        )
    finally:
        # 清理临时文件
        if temp_zip.exists():
            temp_zip.unlink()


@app.post("/admin/upload-batch")
async def upload_batch(
    request: Request,
    files: List[UploadFile] = File(...),
    source_type: str = Form(default="opensource"),
    _: bool = Depends(require_auth)
):
    """Upload multiple skill ZIP files (batch upload)."""
    import tempfile

    # Get user_id from session at the beginning
    user_id = request.session.get("user_id")

    if not files:
        return templates.TemplateResponse("admin_upload.html", {
            "request": request,
            "success": None,
            "error": "No files selected"
        })

    results = {
        "success": [],
        "failed": []
    }

    # Get source_type for each file from form data (source_type_0, source_type_1, etc.)
    form_data = await request.form()

    for idx, file in enumerate(files):
        # Skip non-ZIP files
        if not file.filename or not file.filename.endswith('.zip'):
            results["failed"].append({"file": file.filename, "error": "Not a ZIP file"})
            continue

        # Get source type for this specific file (default to 'opensource' or use global source_type)
        file_source_type = form_data.get(f"source_type_{idx}", source_type)

        temp_dir = tempfile.mkdtemp()
        temp_zip = Path(temp_dir) / "upload.zip"

        try:
            # Save uploaded file
            with open(temp_zip, "wb") as f:
                shutil.copyfileobj(file.file, f)

            # Validate the ZIP file
            is_valid, metadata = validate_skill_zip(temp_zip, default_author=user_id)

            if not is_valid:
                results["failed"].append({
                    "file": file.filename,
                    "error": metadata.get('error', 'Unknown error')
                })
                continue

            # Check if skill with same name already exists
            from database import get_skill_by_name
            skill_name = metadata["name"]
            skill_version = metadata.get("version", "1.0.0")
            target_filename = f"{skill_name}-{skill_version}.zip"

            existing_skill = get_skill_by_name(skill_name)
            if existing_skill:
                if existing_skill["uploader_id"] != user_id:
                    results["failed"].append({
                        "file": file.filename,
                        "error": f"Skill {skill_name} is owned by another user"
                    })
                    continue
                # Skill exists and belongs to user - this is an update, continue

            # Save the skill ZIP to pending directory for review
            target_path = PENDING_DIR / target_filename
            shutil.copy(temp_zip, target_path)

            # Create database record with source_type
            create_skill_record(
                skill_name=skill_name,
                version=skill_version,
                filename=target_filename,
                uploader_id=user_id,
                status='pending',  # Batch uploads also require admin approval
                source_type=file_source_type
            )

            results["success"].append({
                "name": metadata["name"],
                "version": metadata["version"]
            })

        except Exception as e:
            results["failed"].append({"file": file.filename, "error": str(e)})

        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    # Build result message
    success_count = len(results["success"])
    failed_count = len(results["failed"])

    if success_count > 0 and failed_count == 0:
        success_msg = f"Successfully uploaded {success_count} skill(s): " + \
                      ", ".join([f"{s['name']}@{s['version']}" for s in results["success"]])
        return templates.TemplateResponse("admin_upload.html", {
            "request": request,
            "success": success_msg,
            "error": None
        })
    elif success_count > 0 and failed_count > 0:
        success_msg = f"Uploaded {success_count} skill(s), {failed_count} failed"
        error_msg = "; ".join([f"{f['file']}: {f['error']}" for f in results["failed"]])
        return templates.TemplateResponse("admin_upload.html", {
            "request": request,
            "success": success_msg,
            "error": error_msg
        })
    else:
        error_msg = "All uploads failed: " + \
                    "; ".join([f"{f['file']}: {f['error']}" for f in results["failed"]])
        return templates.TemplateResponse("admin_upload.html", {
            "request": request,
            "success": None,
            "error": error_msg
        })


@app.delete("/admin/plugins/{filename}")
async def delete_plugin(
    filename: str,
    _: bool = Depends(require_admin)
):
    """Delete a plugin (admin only)."""
    file_path = PLUGINS_DIR / filename

    if not file_path.exists():
        raise HTTPException(404, "Plugin not found")

    file_path.unlink()

    return {"success": True, "message": f"Deleted {filename}"}




@app.post("/api/batch-download")
async def batch_download(request: Request):
    """Generate a batch download package (ZIP) for selected skills.

    Each skill is packaged with installer scripts (install.bat and install.sh).
    Also includes install-all scripts at the root level for installing all skills at once.
    """
    import zipfile
    from io import BytesIO

    try:
        # Parse request data
        data = await request.json()
        selected_filenames = data.get("filenames", [])

        if not selected_filenames:
            raise HTTPException(400, "No skills selected")

        # Create ZIP in memory - packaging original ZIP files only
        zip_buffer = BytesIO()

        added_count = 0

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for filename in selected_filenames:
                target_zip = PLUGINS_DIR / filename

                if not target_zip.exists():
                    logger.warning(f"ZIP file not found: {filename}", extra={"filename": filename})
                    continue

                # Add original ZIP file directly to batch package
                zf.write(target_zip, filename)
                added_count += 1

            if added_count == 0:
                raise HTTPException(404, "No valid skill files found")

        # Get ZIP data - returning original ZIP files only
        zip_data = zip_buffer.getvalue()

        # Return response with ZIP data
        return Response(
            content=zip_data,
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename=skills-batch-{datetime.now().strftime('%Y%m%d-%H%M%S')}.zip"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Failed to create batch package: {str(e)}")


def generate_install_all_bat(skills: list) -> str:
    """Generate Windows batch script to install all skills."""
    skill_list = '\n'.join([f'echo   - {s["name"]} v{s["version"]}' for s in skills])

    # Build skill installation block
    install_blocks = []
    for skill in skills:
        skill_name = skill['name']
        skill_dir = skill['dir']
        install_blocks.append(f"""
echo 正在安装 {skill_name}...
call :install_skill "{skill_name}" "{skill_dir}"
if %INSTALL_ERRORS% gtr 0 (
    echo {skill_name} 安装失败
)
echo.""")

    install_block_str = '\n'.join(install_blocks)

    content = f"""@echo off
chcp 65001 >nul
title Claude Code Skills 批量安装程序

setlocal EnableDelayedExpansion

echo ========================================
echo   Claude Code Skills 批量安装程序
echo ========================================
echo.
echo 将安装 {len(skills)} 个技能:
{skill_list}
echo.

:: ====================
:: 步骤1: 检测 CLAUDE_SKILLS_PATH
:: ====================
if defined CLAUDE_SKILLS_PATH (
    if exist "%CLAUDE_SKILLS_PATH%" (
        set SKILLS_DIR=%CLAUDE_SKILLS_PATH%
        echo [OK] 从环境变量检测到 Skills 目录: %SKILLS_DIR%
        goto :start_install
    )
)

:: ====================
:: 步骤2: 检测默认路径
:: ====================
set DEFAULT_PATH=%USERPROFILE%\\.claude\\skills
if exist "%DEFAULT_PATH%" (
    set SKILLS_DIR=%DEFAULT_PATH%
    echo [OK] 从默认位置检测到 Skills 目录: %SKILLS_DIR%
    goto :start_install
)

:: ====================
:: 步骤3: 交互式输入
:: ====================
echo.
echo [WARN] 未自动检测到 Claude Code Skills 目录
echo.
echo 常见位置:
echo   - Windows: %%USERPROFILE%%\\.claude\\skills\
echo.

:input_loop
set /p SKILLS_DIR="请输入 Skills 目录完整路径: "

:: 去除首尾空格
for /f "tokens=*" %%a in ("%SKILLS_DIR%") do set SKILLS_DIR=%%a

:: 检查路径是否存在
if not exist "%SKILLS_DIR%" (
    echo [ERR] 路径不存在: %SKILLS_DIR%
    choice /c YN /n /m "是否创建此目录? (Y/N) "
    if errorlevel 2 goto :input_loop
    if errorlevel 1 (
        mkdir "%SKILLS_DIR%" 2>nul
        if errorlevel 1 (
            echo [ERR] 创建目录失败，请检查权限
            goto :input_loop
        )
        echo [OK] 已创建目录: %SKILLS_DIR%
    )
)

:start_install
echo.
echo 目标安装路径: %SKILLS_DIR%
echo.
pause
echo.
echo 开始安装...
echo.

set INSTALL_ERRORS=0
set SCRIPT_DIR=%~dp0

{install_block_str}

echo ========================================
if %INSTALL_ERRORS% == 0 (
    echo 所有技能安装完成!
) else (
    echo 安装完成，部分技能可能未成功 (错误数: %INSTALL_ERRORS%)
)
echo ========================================
echo.
echo 请重启 Claude Code 以加载新技能
echo.
pause
exit /b 0

:: ====================
:: 安装单个技能的子程序
:: ====================
:install_skill
set SKILL_NAME=%~1
set SKILL_DIR=%~2
set SOURCE_PATH=%SCRIPT_DIR%%SKILL_NAME%
set TARGET_PATH=%SKILLS_DIR%\\%SKILL_DIR%

:: 检查源目录是否存在
if not exist "%SOURCE_PATH%" (
    echo [ERR] 未找到技能目录: %SKILL_NAME%
    set /a INSTALL_ERRORS+=1
    exit /b 1
)

:: 检查目标是否已存在
if exist "%TARGET_PATH%" (
    echo [WARN] 技能已存在: %SKILL_DIR%
    choice /c OSB /n /m "请选择 [O]覆盖 [S]跳过 [B]备份: "

    if errorlevel 3 (
        :: 备份
        set BACKUP_NAME=%SKILL_DIR%.backup.%date:~0,4%%date:~5,2%%date:~8,2%-%time:~0,2%%time:~3,2%%time:~6,2%
        set BACKUP_NAME=!BACKUP_NAME: =0!
        rename "%TARGET_PATH%" "!BACKUP_NAME!" >nul 2>&1
        if errorlevel 1 (
            echo [ERR] 备份失败
            set /a INSTALL_ERRORS+=1
            exit /b 1
        )
        echo [OK] 已备份旧版本为: !BACKUP_NAME!
    )

    if errorlevel 2 (
        :: 跳过
        echo [WARN] 已跳过 %SKILL_NAME%
        exit /b 0
    )

    if errorlevel 1 (
        :: 覆盖
        rmdir /s /q "%TARGET_PATH%" >nul 2>&1
        if errorlevel 1 (
            echo [ERR] 删除旧版本失败
            set /a INSTALL_ERRORS+=1
            exit /b 1
        )
        echo [OK] 已删除旧版本
    )
)

:: 执行复制
xcopy "%SOURCE_PATH%" "%TARGET_PATH%\" /s /e /i /q >nul 2>&1
if errorlevel 1 (
    echo [ERR] 复制失败: %SKILL_NAME%
    set /a INSTALL_ERRORS+=1
    exit /b 1
)

echo [OK] %SKILL_NAME% 安装成功
exit /b 0
"""
    return "\ufeff" + content


def generate_install_all_sh(skills: list) -> str:
    """Generate Linux/macOS shell script to install all skills."""
    skill_list = '\n'.join([f'echo "  - {s["name"]} v{s["version"]}"' for s in skills])

    # Build skill installation block
    install_blocks = []
    for skill in skills:
        skill_name = skill['name']
        skill_dir = skill['dir']
        install_blocks.append(f'''echo "安装 {skill_name}..."
install_skill "{skill_name}" "{skill_dir}"
if [ $? -ne 0 ]; then
    INSTALL_ERRORS=$((INSTALL_ERRORS + 1))
fi''')

    install_block_str = '\n'.join(install_blocks)

    return f"""#!/bin/bash

# Claude Code Skills 批量安装脚本

# Colors
GREEN='\\033[0;32m'
YELLOW='\\033[1;33m'
RED='\\033[0;31m'
CYAN='\\033[0;36m'
NC='\\033[0m'

echo ''
echo -e "${{CYAN}}========================================${{NC}}"
echo -e "${{CYAN}}  Claude Code Skills 批量安装程序${{NC}}"
echo -e "${{CYAN}}========================================${{NC}}"
echo ''
echo "将安装 {len(skills)} 个技能:"
{skill_list}
echo ''

# ====================
# 步骤1: 检测 CLAUDE_SKILLS_PATH
# ====================
if [ -n "$CLAUDE_SKILLS_PATH" ] && [ -d "$CLAUDE_SKILLS_PATH" ]; then
    SKILLS_DIR="$CLAUDE_SKILLS_PATH"
    echo -e "${{GREEN}}[✓]${{NC}} 从环境变量检测到 Skills 目录: $SKILLS_DIR"
elif [ -d "$HOME/.claude/skills" ]; then
    SKILLS_DIR="$HOME/.claude/skills"
    echo -e "${{GREEN}}[✓]${{NC}} 从默认位置检测到 Skills 目录: $SKILLS_DIR"
else
    echo -e "${{YELLOW}}[!]${{NC}} 未自动检测到 Claude Code Skills 目录"
    echo ''
    echo '常见位置:'
    echo '  - macOS/Linux: ~/.claude/skills/'
    echo ''

    while true; do
        read -rp "请输入 Skills 目录完整路径: " SKILLS_DIR
        SKILLS_DIR="${{SKILLS_DIR/#\\~/$HOME}}"
        SKILLS_DIR="$(echo "$SKILLS_DIR" | sed 's/^[[:space:]]*//;s/[[:space:]]*$//')"

        if [ -z "$SKILLS_DIR" ]; then
            echo -e "${{RED}}[✗]${{NC}} 路径不能为空"
            continue
        fi

        if [ ! -d "$SKILLS_DIR" ]; then
            echo -e "${{YELLOW}}[!]${{NC}} 路径不存在: $SKILLS_DIR"
            read -rp "是否创建此目录? (Y/N) " create_choice
            if [[ $create_choice =~ ^[Yy]$ ]]; then
                if mkdir -p "$SKILLS_DIR"; then
                    echo -e "${{GREEN}}[✓]${{NC}} 已创建目录: $SKILLS_DIR"
                    break
                else
                    echo -e "${{RED}}[✗]${{NC}} 创建目录失败，请检查权限"
                    continue
                fi
            else
                continue
            fi
        fi
        break
    done
fi

echo ''
echo "目标安装路径: $SKILLS_DIR"
echo ''
read -rp "按 Enter 键开始安装..."
echo ''
echo "开始安装..."
echo ''

INSTALL_ERRORS=0
SCRIPT_DIR="$(cd "$(dirname "${{BASH_SOURCE[0]}}")" && pwd)"

# 安装单个技能的函数
install_skill() {{
    local SKILL_NAME="$1"
    local SKILL_DIR="$2"
    local SOURCE_PATH="$SCRIPT_DIR/$SKILL_NAME"
    local TARGET_PATH="$SKILLS_DIR/$SKILL_DIR"

    # 检查源目录
    if [ ! -d "$SOURCE_PATH" ]; then
        echo -e "${{RED}}[✗]${{NC}} 未找到技能目录: $SKILL_NAME"
        return 1
    fi

    # 检查目标是否已存在
    if [ -d "$TARGET_PATH" ]; then
        echo -e "${{YELLOW}}[!]${{NC}} 技能已存在: $SKILL_DIR"
        echo "请选择: [O]覆盖 [S]跳过 [B]备份"
        read -rp "> " choice

        case "$choice" in
            [Bb])
                local BACKUP_NAME="${{SKILL_DIR}}.backup.$(date +%Y%m%d-%H%M%S)"
                if mv "$TARGET_PATH" "$SKILLS_DIR/$BACKUP_NAME"; then
                    echo -e "${{GREEN}}[✓]${{NC}} 已备份旧版本为: $BACKUP_NAME"
                else
                    echo -e "${{RED}}[✗]${{NC}} 备份失败"
                    return 1
                fi
                ;;
            [Ss])
                echo -e "${{YELLOW}}[!]${{NC}} 已跳过 $SKILL_NAME"
                return 0
                ;;
            [Oo]|*)
                rm -rf "$TARGET_PATH"
                if [ $? -eq 0 ]; then
                    echo -e "${{GREEN}}[✓]${{NC}} 已删除旧版本"
                else
                    echo -e "${{RED}}[✗]${{NC}} 删除旧版本失败"
                    return 1
                fi
                ;;
        esac
    fi

    # 执行复制
    if cp -R "$SOURCE_PATH" "$TARGET_PATH"; then
        echo -e "${{GREEN}}[✓]${{NC}} $SKILL_NAME 安装成功"
        return 0
    else
        echo -e "${{RED}}[✗]${{NC}} 复制失败: $SKILL_NAME"
        return 1
    fi
}}

{install_block_str}

echo ''
echo -e "${{CYAN}}========================================${{NC}}"
if [ $INSTALL_ERRORS -eq 0 ]; then
    echo -e "${{GREEN}}所有技能安装完成!${{NC}}"
else
    echo -e "${{YELLOW}}安装完成，部分技能可能未成功 (错误数: $INSTALL_ERRORS)${{NC}}"
fi
echo -e "${{CYAN}}========================================${{NC}}"
echo ''
echo "请重启 Claude Code 以加载新技能"
echo ''
read -rp "按 Enter 键退出..."
"""


def generate_batch_readme(skills: list) -> str:
    """Generate README for batch download package."""
    skill_list = '\n'.join([f"  - {s['name']} v{s['version']}" for s in skills])

    return f"""Claude Code Skills 批量安装包
{'=' * 50}

包含技能 ({len(skills)}个):
{skill_list}

一键安装所有技能:
-----------------
Windows:
  1. 解压此 ZIP 文件
  2. 双击运行 install-all.bat
  3. 按提示完成所有技能的安装

Linux / macOS:
  1. 解压此 ZIP 文件
  2. 打开终端，cd 到解压目录
  3. 运行: chmod +x install-all.sh \u0026\u0026 ./install-all.sh
  4. 按提示完成所有技能的安装

目录结构:
---------
skills-batch/
├── install-all.bat      # Windows 批量安装脚本（安装所有技能）
├── install-all.sh       # Linux/macOS 批量安装脚本（安装所有技能）
├── README.txt           # 本文件
├── skill-name-1/        # 第一个技能目录
│   ├── package.json
│   └── ...
├── skill-name-2/        # 第二个技能目录
│   └── ...
└── ...

注意事项:
---------
1. install-all 脚本会安装所有包含的技能
2. 安装脚本会自动检测 Claude Code Skills 目录
3. 支持设置 CLAUDE_SKILLS_PATH 环境变量指定自定义路径
4. 如果技能已存在，会提示选择：覆盖/跳过/备份
5. 安装完成后需要重启 Claude Code 以加载新技能

单独安装某个技能:
-----------------
如需单独安装某个技能，可以直接将该技能的目录复制到 Claude Code Skills 目录下:
  - Windows: %USERPROFILE%\\.claude\\skills\
  - Linux/Mac: ~/.claude/skills/

或者下载单个技能的安装包（包含独立安装脚本）。

"""


# Health check endpoint for third-party monitoring
@app.get("/api/health")
async def health_check():
    """Health check endpoint for monitoring services.

    Returns:
        Service status information including:
        - status: "healthy" or "unhealthy"
        - timestamp: Current server time
        - version: API version
        - uptime: Service uptime (if available)
    """
    import time

    # Basic health checks
    health_status = {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "version": "1.0.0",
        "service": "Skill Registry"
    }

    # Check if plugins directory is accessible
    try:
        if not PLUGINS_DIR.exists():
            health_status["status"] = "unhealthy"
            health_status["error"] = "Plugins directory not accessible"
    except Exception as e:
        health_status["status"] = "unhealthy"
        health_status["error"] = str(e)

    # Return appropriate status code
    if health_status["status"] == "healthy":
        return health_status
    else:
        raise HTTPException(status_code=503, detail=health_status)


# Statistics APIs
@app.get("/api/stats/top")
async def api_stats_top(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)")
):
    """Get download statistics with rankings."""
    try:
        # Parse dates
        start = date.fromisoformat(start_date) if start_date else None
        end = date.fromisoformat(end_date) if end_date else None

        # Get plugins for author mapping
        plugins = scan_plugins()

        # Get stats with author info
        stats = get_stats_with_author(plugins, start, end)

        return {
            "period": {
                "start_date": start_date or "all-time",
                "end_date": end_date or "all-time"
            },
            "total_downloads": stats["total_downloads"],
            "rankings": stats["rankings"]
        }

    except ValueError as e:
        raise HTTPException(400, f"Invalid date format: {e}")
    except Exception as e:
        raise HTTPException(500, f"Failed to get stats: {e}")


@app.get("/api/stats/uploads")
async def api_stats_uploads():
    """Get upload statistics for dashboard (public, no auth required).

    Returns:
        - total_skills: Total number of uploaded skills
        - this_month: Number of skills uploaded this month
        - last_month: Number of skills uploaded last month
        - top_uploaders: List of top 10 uploaders with username and count
    """
    try:
        stats = get_upload_stats()
        return {
            "success": True,
            "data": stats
        }
    except Exception as e:
        raise HTTPException(500, f"Failed to get upload stats: {e}")


@app.get("/api/stats/export")
async def api_stats_export(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)")
):
    """Export download statistics as Excel file."""
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # Parse dates
        start = date.fromisoformat(start_date) if start_date else None
        end = date.fromisoformat(end_date) if end_date else None

        # Get plugins for author mapping
        plugins = scan_plugins()

        # Get stats with author info
        stats = get_stats_with_author(plugins, start, end)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Download Statistics"

        # Header row
        headers = ["排名", "Skill 名称", "作者", "下载次数"]
        ws.append(headers)

        # Style header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows
        for idx, ranking in enumerate(stats["rankings"], 1):
            ws.append([
                idx,
                ranking["skill_name"],
                ranking["author"],
                ranking["downloads"]
            ])

        # Style data rows
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in range(2, len(stats["rankings"]) + 2):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if col in [1, 4] else "left", vertical="center")

        # Adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename
        period_str = f"{start_date or 'all'}_to_{end_date or 'all'}"
        filename = f"download_stats_{period_str}.xlsx"

        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except ImportError:
        raise HTTPException(500, "Excel export requires openpyxl: pip install openpyxl")
    except ValueError as e:
        raise HTTPException(400, f"Invalid date format: {e}")
    except Exception as e:
        raise HTTPException(500, f"Failed to export stats: {e}")


@app.get("/stats", response_class=HTMLResponse)
async def stats_page(request: Request):
    """Display download statistics page (public access)."""
    return templates.TemplateResponse("stats.html", {
        "request": request
    })


@app.get("/skill/{skill_name}", response_class=HTMLResponse)
async def skill_detail_page(request: Request, skill_name: str, version: str = None, compare: str = None):
    """Display skill detail page with Skill.md content."""
    # Check if user is authenticated
    user_id = request.session.get("user_id")
    if not user_id:
        return RedirectResponse(url="/login", status_code=302)

    # Try to get skill from database first
    from database import get_skill_by_name
    skill = get_skill_by_name(skill_name)

    # Get all versions of this skill
    from db.repositories import SkillRepository
    all_versions = SkillRepository.get_versions(skill_name)

    # Get all approved versions for display
    approved_versions = [v for v in all_versions if v.get("status") == "approved"]

    skill_zip = None
    real_skill_name = skill_name
    selected_version = version

    # Function to find skill zip by version
    def find_skill_zip(name: str, ver: str = None):
        if ver:
            # Try exact version match: skill-name-version.zip
            exact = PLUGINS_DIR / f"{name}-{ver}.zip"
            if exact.exists():
                return exact
        # Try default or latest
        if skill and skill["status"] == "approved" and skill["is_active"]:
            zip_path = PLUGINS_DIR / skill["filename"]
            if zip_path.exists():
                return zip_path
        # Try exact match first
        exact_match = PLUGINS_DIR / f"{name}.zip"
        if exact_match.exists():
            return exact_match
        # Try pattern match (skill-name-*.zip)
        matching_zips = list(PLUGINS_DIR.glob(f"{name}-*.zip"))
        if matching_zips:
            return matching_zips[0]
        return None

    skill_zip = find_skill_zip(skill_name, version)

    if not skill_zip:
        # No skill found or file not found
        skill_zip = find_skill_zip(skill_name)
        if not skill_zip:
            raise HTTPException(status_code=404, detail=f"Skill not found: {skill_name}")

    # Extract metadata from the skill
    metadata = extract_metadata(skill_zip.name)

    # Get the real skill name from metadata
    real_skill_name = metadata.get("name", skill_name) if metadata else skill_name

    # Get download count from database
    from database import get_download_stats
    stats = get_download_stats()
    download_count = 0
    for ranking in stats["rankings"]:
        if ranking["skill_name"] == real_skill_name:
            download_count = ranking["downloads"]
            break

    # Get author from metadata or database
    author = "Unknown"
    if metadata and "metadata" in metadata:
        author_meta = metadata["metadata"].get("author", "")
        if isinstance(author_meta, dict):
            author = author_meta.get("name", "")
        else:
            author = str(author_meta) if author_meta else ""

    # If author not in metadata, try to get from database uploader
    if not author or author == "Unknown":
        skill_record = get_skill_by_name(real_skill_name)
        if skill_record and skill_record.get("uploader_id"):
            uploader = get_user_by_id(skill_record["uploader_id"])
            if uploader:
                author = uploader.get("employee_id", "Unknown")

    # Get version
    version = metadata.get("version", "1.0.0") if metadata else "1.0.0"

    # Get updated_at from file modification time
    updated_at = datetime.fromtimestamp(skill_zip.stat().st_mtime).strftime("%Y-%m-%d")

    # Get current user
    user = get_current_user(request)

    # Get Gitea repo URL for Agent installation
    gitea_repo_url = os.getenv("GITEA_REPO_URL", "")

    # Build skill directory path for Gitea (format: {skill_name}/SKILL.md)
    skill_dir = real_skill_name

    # Get version for display (use selected version or default)
    display_version = selected_version if selected_version else version

    # Prepare version list for template
    version_list = []
    for v in approved_versions:
        v_info = {
            "version": v.get("version", "unknown"),
            "is_default": v.get("is_default_version", 0) == 1,
            "is_active": v.get("is_active", 1) == 1,
            "created_at": v.get("created_at", "").strftime("%Y-%m-%d") if hasattr(v.get("created_at"), "strftime") else str(v.get("created_at", ""))
        }
        version_list.append(v_info)

    # Get compare version content if specified
    compare_content = None
    compare_version_display = None
    if compare:
        compare_zip = find_skill_zip(skill_name, compare)
        if compare_zip:
            import zipfile
            try:
                with zipfile.ZipFile(compare_zip, 'r') as zf:
                    # Find SKILL.md (may be in root or subdirectory)
                    skill_md_paths = [name for name in zf.namelist()
                                     if 'SKILL.md' in name or name.endswith('SKILL.md')]
                    if skill_md_paths:
                        compare_content = zf.read(skill_md_paths[0]).decode('utf-8')
            except Exception as e:
                logger.warning(f"Failed to extract compare content: {e}")
            compare_version_display = compare

    return templates.TemplateResponse("skill_detail.html", {
        "request": request,
        "skill_name": real_skill_name,
        "author": author,
        "download_count": download_count,
        "version": display_version,
        "updated_at": updated_at,
        "download_url": f"/plugins/{skill_zip.name}",
        "filename": skill_zip.name,
        "request_url": str(request.base_url).rstrip("/"),
        "user": user,
        "gitea_repo_url": gitea_repo_url,
        "skill_dir": skill_dir,
        "version_list": version_list,
        "selected_version": selected_version,
        "compare": compare,
        "compare_content": compare_content,
        "compare_version_display": compare_version_display
    })


@app.get("/skill/{skill_name}/skill.md", response_class=PlainTextResponse)
async def get_skill_md_file(skill_name: str):
    """Get SKILL.md file for Agent installation.

    Returns the SKILL.md file content as plain text for curl download.
    """
    logger.debug(f"SKILL.md requested for: '{skill_name}'")

    # Find the skill ZIP file
    skill_zip = None

    # Try exact match first
    exact_match = PLUGINS_DIR / f"{skill_name}.zip"
    if exact_match.exists():
        skill_zip = exact_match
    else:
        # Try pattern match (skill-name-*.zip)
        matching_zips = list(PLUGINS_DIR.glob(f"{skill_name}-*.zip"))
        if matching_zips:
            skill_zip = matching_zips[0]

    if not skill_zip:
        raise HTTPException(status_code=404, detail=f"Skill not found: {skill_name}")

    try:
        import zipfile

        with zipfile.ZipFile(skill_zip, 'r') as zf:
            # Find SKILL.md (may be in root or subdirectory)
            skill_md_paths = [name for name in zf.namelist()
                             if 'SKILL.md' in name or name.endswith('SKILL.md')]

            if not skill_md_paths:
                raise HTTPException(status_code=404, detail=f"SKILL.md not found in {skill_name}")

            # Read and return SKILL.md content
            skill_md_path = skill_md_paths[0]
            content = zf.read(skill_md_path).decode('utf-8')

            return PlainTextResponse(content=content, media_type="text/markdown")

    except zipfile.BadZipFile:
        raise HTTPException(status_code=500, detail="Invalid ZIP file")
    except Exception as e:
        logger.error(f"Error reading SKILL.md: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/skill/{skill_name}/content")
async def get_skill_content_api(skill_name: str, version: str = None):
    """Get Skill.md content for a skill.

    Returns the complete SKILL.md file content (including YAML frontmatter).
    """
    logger.debug(f"API called with skill_name: '{skill_name}', version: '{version}'", extra={"skill_name": skill_name, "version": version})

    # Find the skill ZIP file by version
    def find_zip(name: str, ver: str = None):
        if ver:
            zip_path = PLUGINS_DIR / f"{name}-{ver}.zip"
            if zip_path.exists():
                return zip_path
        # Fallback to default behavior
        zip_path = PLUGINS_DIR / f"{name}.zip"
        if zip_path.exists():
            return zip_path
        matching_zips = list(PLUGINS_DIR.glob(f"{name}-*.zip"))
        if matching_zips:
            return matching_zips[0]
        return None

    skill_zip = find_zip(skill_name, version)

    if not skill_zip:
        raise HTTPException(status_code=404, detail=f"Skill ZIP file not found for: {skill_name}")

    try:
        import zipfile

        with zipfile.ZipFile(skill_zip, 'r') as zf:
            # Find SKILL.md (may be in root or subdirectory)
            skill_md_paths = [name for name in zf.namelist()
                             if 'SKILL.md' in name or name.endswith('SKILL.md')]

            logger.debug(f"Found SKILL.md files: {skill_md_paths}", extra={"skill_name": skill_name, "skill_md_paths": skill_md_paths})

            if not skill_md_paths:
                return {"content": None}

            # Read SKILL.md content
            skill_md_path = skill_md_paths[0]
            content = zf.read(skill_md_path).decode('utf-8')

            logger.debug(f"SKILL.md content length: {len(content)} bytes", extra={"skill_name": skill_name, "content_length": len(content)})

            # Return the complete SKILL.md content (including YAML frontmatter)
            return {"content": content}

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to read skill content: {str(e)}"
        )


@app.get("/api/admin/stats")
async def api_admin_stats(
    _: bool = Depends(require_admin)
):
    """Get admin statistics (admin only).

    Returns comprehensive statistics about the registry including:
    - Total users count
    - Pending skills count
    - Approved skills count
    - Today's downloads count
    - Top 10 skills by downloads
    - Top 10 users by downloads
    """
    try:
        # Get counts
        total_users = get_total_users_count()
        pending_skills = get_skills_count_by_status("pending")
        approved_skills = get_skills_count_by_status("approved")
        today_downloads = get_today_downloads_count()

        # Get top rankings
        top_skills = get_top_skills_by_downloads(10)
        top_users = get_top_users_by_downloads(10)

        return {
            "success": True,
            "data": {
                "total_users": total_users,
                "pending_skills": pending_skills,
                "approved_skills": approved_skills,
                "today_downloads": today_downloads,
                "top_skills": top_skills,
                "top_users": top_users
            }
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch admin statistics: {str(e)}"
        )


@app.put("/api/admin/skills/{skill_id}/source-type")
async def api_update_skill_source_type(
    skill_id: int,
    request: Request,
    _: bool = Depends(require_admin)
):
    """Update the source_type of a skill (admin only).

    Expects JSON body with:
    {
        "source_type": "opensource" | "icsl" | "huawei"
    }
    """
    try:
        # Get current user
        reviewer_id = request.session.get("user_id")
        if not reviewer_id:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Not authenticated"
            )

        # Parse request body
        data = await request.json()
        source_type = data.get("source_type")

        # Validate source_type
        valid_types = ["opensource", "icsl", "huawei"]
        if source_type not in valid_types:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid source_type. Must be one of: {', '.join(valid_types)}"
            )

        # Get skill record
        skill = get_skill_by_id(skill_id)
        if not skill:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Skill {skill_id} not found"
            )

        # Update source_type in database
        from database import get_connection
        with get_connection() as conn:
            conn.execute(
                """
                UPDATE skills
                SET source_type = %s
                WHERE id = %s
                """,
                (source_type, skill_id)
            )
            conn.commit()

        return {
            "success": True,
            "message": f"Updated source_type for {skill['skill_name']} to {source_type}",
            "skill_id": skill_id,
            "source_type": source_type
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update source_type: {str(e)}"
        )


@app.get("/api/admin/skills")
async def api_get_all_skills(
    status: Optional[str] = None,
    limit: int = Query(100, ge=1, le=500),
    _: bool = Depends(require_admin)
) -> Dict[str, Any]:
    """Get all skills with optional status filter (admin only).

    Args:
        status: Filter by status (pending/approved/rejected)
        limit: Maximum number of skills to return

    Returns:
        List of skills with full details
    """
    try:
        from database import get_connection

        with get_connection() as conn:
            if status:
                rows = conn.execute("""
                    SELECT s.id, s.skill_name, s.version, s.filename, s.status,
                           s.source_type, s.uploaded_at, u.employee_id as uploader_name
                    FROM skills s
                    LEFT JOIN users u ON s.uploader_id = u.id
                    WHERE s.status = %s
                    ORDER BY s.uploaded_at DESC
                    LIMIT %s
                """, (status, limit)).fetchall()
            else:
                rows = conn.execute("""
                    SELECT s.id, s.skill_name, s.version, s.filename, s.status,
                           s.source_type, s.uploaded_at, u.employee_id as uploader_name
                    FROM skills s
                    LEFT JOIN users u ON s.uploader_id = u.id
                    ORDER BY s.uploaded_at DESC
                    LIMIT %s
                """, (limit,)).fetchall()

            return {
                "success": True,
                "data": rows,
                "count": len(rows)
            }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch skills: {str(e)}"
        )


@app.get("/api/admin/gitea-tasks")
async def api_get_gitea_tasks(
    status: Optional[str] = None,
    limit: int = Query(50, ge=1, le=200),
    _: bool = Depends(require_admin)
) -> Dict[str, Any]:
    """Get Gitea push tasks with optional status filter.

    Args:
        status: Filter by status (pending/pushing/success/failed)
        limit: Maximum number of tasks to return

    Returns:
        List of push tasks with skill info
    """
    try:
        from database import get_connection

        with get_connection() as conn:
            if status:
                rows = conn.execute("""
                    SELECT t.*, s.skill_name, s.uploader_id, u.employee_id as uploader_name
                    FROM gitea_push_tasks t
                    JOIN skills s ON t.skill_id = s.id
                    LEFT JOIN users u ON s.uploader_id = u.id
                    WHERE t.status = %s
                    ORDER BY t.created_at DESC
                    LIMIT %s
                """, (status, limit)).fetchall()
            else:
                rows = conn.execute("""
                    SELECT t.*, s.skill_name, s.uploader_id, u.employee_id as uploader_name
                    FROM gitea_push_tasks t
                    JOIN skills s ON t.skill_id = s.id
                    LEFT JOIN users u ON s.uploader_id = u.id
                    ORDER BY t.created_at DESC
                    LIMIT %s
                """, (limit,)).fetchall()

            return {
                "success": True,
                "data": rows,
                "count": len(rows)
            }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch tasks: {str(e)}"
        )


@app.post("/api/admin/gitea-tasks/{task_id}/retry")
async def api_retry_gitea_task(
    task_id: int,
    _: bool = Depends(require_admin)
) -> Dict[str, Any]:
    """Retry a failed Gitea push task.

    Args:
        task_id: ID of the task to retry

    Returns:
        Success message
    """
    try:
        from database import get_connection

        with get_connection() as conn:
            # Check task status
            row = conn.execute("""
                SELECT status, skill_id FROM gitea_push_tasks WHERE id = %s
            """, (task_id,)).fetchone()

            if not row:
                raise HTTPException(status_code=404, detail="Task not found")

            if row["status"] == "success":
                raise HTTPException(status_code=400, detail="Task already succeeded")

            # Reset task to pending
            conn.execute("""
                UPDATE gitea_push_tasks
                SET status = 'pending', error_message = NULL, retry_count = retry_count + 1
                WHERE id = %s
            """, (task_id,))
            conn.commit()

            return {
                "success": True,
                "message": f"Task {task_id} queued for retry"
            }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to retry task: {str(e)}"
        )


@app.post("/api/admin/gitea-push/trigger")
async def api_trigger_gitea_push(
    _: bool = Depends(require_admin)
) -> Dict[str, Any]:
    """Manually trigger Gitea push service to process pending tasks.

    Returns:
        Success message with task count
    """
    try:
        from services.gitea.gitea_push_service import run_push_task

        # Run the push task synchronously
        run_push_task()

        return {
            "success": True,
            "message": "Push triggered successfully"
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to trigger push: {str(e)}"
        )


# ==================== User Management API Endpoints ====================

@app.get("/api/admin/users")
async def api_get_users(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    role: Optional[str] = Query(None, pattern="^(admin|user)$"),
    status_filter: Optional[str] = Query(None, pattern="^(active|disabled)$"),
    search: Optional[str] = Query(None, max_length=50),
    _: bool = Depends(require_admin)
) -> Dict[str, Any]:
    """Get paginated list of users with optional filters.

    Query params:
        page: Page number (default: 1)
        per_page: Users per page (default: 20, max: 100)
        role: Filter by role ('admin' or 'user')
        status_filter: Filter by status ('active' or 'disabled')
        search: Search by employee_id (partial match)

    Returns:
        Paginated user list with total count
    """
    try:
        result = get_users_list(
            page=page,
            per_page=per_page,
            role=role,
            status_filter=status_filter,
            search=search
        )
        return {
            "success": True,
            "data": result
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch users: {str(e)}"
        )


@app.post("/api/admin/users")
async def api_create_user(
    request: Request,
    employee_id: str = Form(..., max_length=50),
    role: str = Form(..., pattern="^(admin|user)$"),
    _: bool = Depends(require_admin)
) -> Dict[str, Any]:
    """Create a new user.

    Form data:
        employee_id: Employee ID (alphanumeric, max 50 chars)
        role: User role ('admin' or 'user')

    Returns:
        Created user data with API key (shown only once)
    """
    try:
        # Validate employee_id format
        if not employee_id or not re.match(r'^[a-zA-Z0-9_-]+$', employee_id):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid employee_id format. Must be alphanumeric with underscores/hyphens only."
            )

        # Check if employee_id already exists
        existing_user = get_user_by_credentials(employee_id, "dummy")
        if existing_user:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"User with employee_id '{employee_id}' already exists"
            )

        # Generate 32-char API key
        api_key = secrets.token_hex(16)

        # Create user
        user_id = create_user(employee_id=employee_id, api_key=api_key, role=role)

        # Get the created user
        user = get_user_by_id(user_id)

        return {
            "success": True,
            "data": {
                "id": user["id"],
                "employee_id": user["employee_id"],
                "role": user["role"],
                "api_key": api_key,  # Only shown once on creation
                "created_at": user["created_at"]
            },
            "message": "User created successfully. Save the API key now as it won't be shown again."
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create user: {str(e)}"
        )


@app.put("/api/admin/users/{user_id}")
async def api_update_user_role(
    user_id: int,
    request: Request,
    role: str = Form(..., pattern="^(admin|user)$"),
    _: bool = Depends(require_admin)
) -> Dict[str, Any]:
    """Update a user's role.

    Form data:
        role: New role ('admin' or 'user')

    Prevents editing own role.
    """
    try:
        # Prevent editing own role
        current_user_id = request.session.get("user_id")
        if current_user_id == user_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Cannot modify your own role"
            )

        # Check if user exists
        user = get_user_by_id(user_id)
        if not user:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="User not found"
            )

        # Update role
        success = update_user_role(user_id, role)
        if not success:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Failed to update user role"
            )

        return {
            "success": True,
            "data": {
                "id": user_id,
                "employee_id": user["employee_id"],
                "role": role
            },
            "message": "User role updated successfully"
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update user role: {str(e)}"
        )


@app.patch("/api/admin/users/{user_id}/disable")
async def api_disable_user(
    user_id: int,
    request: Request,
    _: bool = Depends(require_admin)
) -> Dict[str, Any]:
    """Disable a user.

    Prevents login but keeps the user in the system.
    Prevents disabling self.
    """
    try:
        # Prevent disabling self
        current_user_id = request.session.get("user_id")
        if current_user_id == user_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Cannot disable your own account"
            )

        # Check if user exists
        user = get_user_by_id(user_id)
        if not user:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="User not found"
            )

        # Disable user
        success = disable_user(user_id)
        if not success:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Failed to disable user"
            )

        return {
            "success": True,
            "data": {
                "id": user_id,
                "employee_id": user["employee_id"],
                "status": "disabled"
            },
            "message": "User disabled successfully"
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to disable user: {str(e)}"
        )


@app.delete("/api/admin/users/{user_id}")
async def api_delete_user(
    user_id: int,
    request: Request,
    _: bool = Depends(require_admin)
) -> Dict[str, Any]:
    """Permanently delete a user.

    Only allowed if user has no skills (skills_count == 0).
    Prevents deleting self.
    """
    try:
        # Prevent deleting self
        current_user_id = request.session.get("user_id")
        if current_user_id == user_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Cannot delete your own account"
            )

        # Check if user exists
        user = get_user_by_id(user_id)
        if not user:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="User not found"
            )

        # Check if user has active skills
        skills_count = get_user_skills_count(user_id)
        if skills_count > 0:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Cannot delete user with {skills_count} skill(s). Please remove skills first."
            )

        # Delete user
        success = delete_user(user_id)
        if not success:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Failed to delete user"
            )

        return {
            "success": True,
            "data": {
                "id": user_id,
                "employee_id": user["employee_id"]
            },
            "message": "User deleted successfully"
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete user: {str(e)}"
        )


@app.patch("/api/admin/users/{user_id}/enable")
async def api_enable_user(
    user_id: int,
    _: bool = Depends(require_admin)
) -> Dict[str, Any]:
    """Re-enable a disabled user.

    Sets status back to 'active'.
    """
    try:
        # Check if user exists
        user = get_user_by_id(user_id)
        if not user:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="User not found"
            )

        # Enable user
        success = enable_user(user_id)
        if not success:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Failed to enable user"
            )

        return {
            "success": True,
            "data": {
                "id": user_id,
                "employee_id": user["employee_id"],
                "status": "active"
            },
            "message": "User enabled successfully"
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to enable user: {str(e)}"
        )


@app.post("/api/admin/users/{user_id}/reset-key")
async def api_reset_user_api_key(
    user_id: int,
    _: bool = Depends(require_admin)
) -> Dict[str, Any]:
    """Reset a user's API key.

    Generates a new 32-character API key.
    Old key is immediately invalidated.
    Rate limited to once per 5 minutes per user.
    """
    try:
        # Check rate limit (5 minutes)
        from datetime import timedelta
        rate_limit_minutes = 5
        current_time = datetime.now()

        if user_id in _api_key_reset_times:
            last_reset_time = _api_key_reset_times[user_id]
            time_since_reset = current_time - last_reset_time
            if time_since_reset < timedelta(minutes=rate_limit_minutes):
                raise HTTPException(
                    status_code=status.HTTP_429_TOO_MANY_REQUESTS,
                    detail=f"API key reset too frequently. Please wait {rate_limit_minutes} minutes between resets."
                )

        # Check if user exists
        user = get_user_by_id(user_id)
        if not user:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="User not found"
            )

        # Generate new API key
        new_api_key = secrets.token_hex(16)

        # Reset API key
        success = reset_user_api_key(user_id, new_api_key)
        if not success:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Failed to reset API key"
            )

        # Store reset time after successful reset
        _api_key_reset_times[user_id] = current_time

        return {
            "success": True,
            "data": {
                "id": user_id,
                "employee_id": user["employee_id"],
                "api_key": new_api_key
            },
            "message": "API key reset successfully. Save the new key now as it won't be shown again."
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to reset API key: {str(e)}"
        )


# ==================== API Key Management Routes ====================

@app.get("/api/admin/api-keys")
async def api_get_api_keys(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None, max_length=50),
    status_filter: Optional[str] = Query(None, pattern="^(active|inactive)$"),
    _: bool = Depends(require_admin)
) -> Dict[str, Any]:
    """获取 API Keys 列表（管理员）"""
    try:
        result = get_api_keys_list(
            page=page,
            per_page=per_page,
            search=search,
            status_filter=status_filter
        )
        return {
            "success": True,
            "data": result
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch API keys: {str(e)}"
        )


@app.post("/api/admin/api-keys")
async def api_create_api_key(
    name: str = Form(None, max_length=100),
    rate_limit: int = Form(100, ge=1, le=1000),
    _: bool = Depends(require_admin)
) -> Dict[str, Any]:
    """创建新的 API Key（管理员）- 不绑定用户，用于外部 API 调用"""
    try:
        # 使用固定值 0 表示这是管理员创建的全局 API Key
        admin_user_id = 0

        # 创建 API Key
        api_key_info = create_api_key(admin_user_id, name, rate_limit)
        if not api_key_info:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to create API key"
            )

        return {
            "success": True,
            "data": api_key_info,
            "message": "API Key created successfully. Save the key now as it won't be shown again."
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create API key: {str(e)}"
        )


@app.delete("/api/admin/api-keys/{api_key_id}")
async def api_delete_api_key(
    api_key_id: int,
    _: bool = Depends(require_admin)
) -> Dict[str, Any]:
    """删除 API Key（管理员）"""
    try:
        success = delete_api_key(api_key_id)
        if not success:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="API Key not found"
            )

        return {
            "success": True,
            "message": "API Key deleted successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete API key: {str(e)}"
        )


@app.put("/api/admin/api-keys/{api_key_id}/toggle")
async def api_toggle_api_key(
    api_key_id: int,
    _: bool = Depends(require_admin)
) -> Dict[str, Any]:
    """切换 API Key 状态（启用/禁用）"""
    try:
        result = toggle_api_key_status(api_key_id)
        if not result:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="API Key not found"
            )

        return {
            "success": True,
            "data": result,
            "message": f"API Key {'enabled' if result['is_active'] else 'disabled'} successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to toggle API key: {str(e)}"
        )


@app.get("/api/admin/api-keys/{api_key_id}/stats")
async def api_get_api_key_stats(
    api_key_id: int,
    _: bool = Depends(require_admin)
) -> Dict[str, Any]:
    """获取 API Key 调用统计（管理员）"""
    try:
        stats = get_api_key_stats(api_key_id)
        if not stats:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="API Key not found"
            )

        return {
            "success": True,
            "data": stats
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get API key stats: {str(e)}"
        )


# ==================== UI Routes ====================

@app.get("/admin/gitea-tasks", response_class=HTMLResponse)
async def gitea_tasks_page(request: Request):
    """Display Gitea push tasks status page."""
    return templates.TemplateResponse("gitea_tasks.html", {
        "request": request
    })


# ==================== Route Modules ====================
# Register additional route modules (lower priority than @app routes)

def _register_additional_routes():
    """Register additional route modules."""
    from apps import pages, downloads, notifications, users, keys

    # Set templates for pages router
    pages.set_templates(templates)

    # Include routers - these have lower priority than @app routes
    app.include_router(pages.router, tags=["Pages"])
    app.include_router(downloads.router, tags=["Downloads"])
    app.include_router(notifications.router, tags=["Notifications"])
    app.include_router(users.router, tags=["Users"])
    app.include_router(keys.router, tags=["API Keys"])


# Register additional routes
_register_additional_routes()


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=28000)
